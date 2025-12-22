from ..base.B24 import B24
from ..base.Elastic.ArticleSearchModel import ArticleSearchModel
from .File import File
from .User import User
from .Tag import Tag
from ..base.pSQL.objects.ArticleModel import ArticleModel
from ..base.pSQL.objects.LikesModel import LikesModel
from ..base.pSQL.objects.ViewsModel import ViewsModel
from ..services.Idea import Idea
from ..services.LogsMaker import LogsMaker

import re
import json
import datetime
import asyncio
import types

from fastapi import APIRouter, Body, Request
import os
from dotenv import load_dotenv

from fastapi import Depends
from sqlalchemy.ext.asyncio import AsyncSession
from ..base.pSQL.objects.App import get_async_db

from openpyxl import Workbook
import io
from fastapi.responses import StreamingResponse
from urllib.parse import quote

load_dotenv()

DOMAIN = os.getenv('HOST')

article_router = APIRouter(prefix="/article")


def make_date_valid(date):
    if date is not None:
        if isinstance(date, str):
            if '-' in date:  
                try:
                    # return datetime.datetime.strptime(date, '%d.%m.%Y %H:%M:%S')
                    return datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
                except:
                    # return datetime.datetime.strptime(date, '%d.%m.%Y %H:%M:%S')
                    # return datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
                    return datetime.datetime.strptime(date, '%Y-%m-%d')
            elif '.' in date:
                try:
                    return datetime.datetime.strptime(date, '%d.%m.%Y %H:%M:%S')
                    # return datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S')
                except:
                    # return datetime.datetime.strptime(date, '%d.%m.%Y %H:%M:%S')
                    return datetime.datetime.strptime(date, '%d.%m.%Y')
                    # return datetime.datetime.strptime(date, '%Y-%m-%d %H:%M:%S')

    else:
        return None


def take_value(PROPERTY: dict | list | str):
    if type(PROPERTY) == type(dict()):
        return list(PROPERTY.values())[0]
    elif type(PROPERTY) == type(list()):
        return PROPERTY[0]
    else:
        return None


def dict_to_indirect_data(data, property_value_dict):
    res = dict()
    for key in property_value_dict.keys():
        if key in data:
            res[property_value_dict[key]] = take_value(data[key])
    return res


def extract_user_data(html_string):
    pattern = r'<div[^>]*>\{\{user id=(\d+);text=([^}]+)\}\}</div>'
    matches = re.findall(pattern, html_string)
    result = [{'id': int(match[0]), 'text': match[1]} for match in matches]
    return result


class Article:
    def __init__(self, id=0, section_id=0):
        self.id = id
        self.section_id = section_id

        # кастомный прогрессбар
        self.logg = LogsMaker()

    async def find(self, inf_id, art_id, property):
        return B24().find(inf_id, art_id, property)

    async def get_inf(self):
        data = B24().getInfoBlock(self.section_id)

        if hasattr(data, '_asyncio_future_blocking'):  # Это Task
            data = await data

        return data

    async def make_valid_article(self, data, session):
        '''
        ! Добавить статью и стандартизировать данные
        '''

        self.id = int(data['ID'])

        if "PREVIEW_TEXT" in data:
            preview = data['PREVIEW_TEXT']
            # data.pop('PREVIEW_TEXT')
        elif "PROPERTY_1009" in data:
            preview = list(data['PROPERTY_1009'].values())[0]
            # data.pop('PROPERTY_1009')
        elif "PROPERTY_341" in data:
            preview = list(data['PROPERTY_341'].values())[0]
            # data.pop('PROPERTY_341')
        elif "PROPERTY_290" in data:
            preview = list(data['PROPERTY_290'].values())[0]
            # data.pop('PROPERTY_290')
        elif "PROPERTY_356" in data:
            preview = list(data['PROPERTY_356'].values())[0]
            # data.pop('PROPERTY_356')
        elif "PROPERTY_488" in data:
            preview = list(data['PROPERTY_488'].values())[0]
            # data.pop('PROPERTY_488')
        elif "PROPERTY_1127" in data:
            preview = list(data['PROPERTY_1127'].values())[0]
            # data.pop('PROPERTY_1127')
        elif "PROPERTY_677" in data:
            preview = list(data['PROPERTY_677'].values())[0]
        else:
            preview = None

        content_type = None
        if "CONTENT_TEXT" in data:
            content = data['CONTENT_TEXT']
            # data.pop('CONTENT_TEXT')
        elif "TEXT" in data:
            content = data['TEXT']
            # data.pop('TEXT')
        elif "DETAIL_TEXT" in data:
            content = data['DETAIL_TEXT']
            # data.pop('DETAIL_TEXT')
        elif "PROPERTY_365" in data:
            content = list(data['PROPERTY_365'].values())[0]
            # data.pop('PROPERTY_365')
        elif "PROPERTY_374" in data:
            content = list(data['PROPERTY_374'].values())[0]["TEXT"]
            content_type = list(data['PROPERTY_374'].values())[0]["TYPE"]

        else:
            keys = ["PROPERTY_1239", "PROPERTY_457", "PROPERTY_477", "PROPERTY_340", "PROPERTY_291", "PROPERTY_358",
                    "PROPERTY_1034", "PROPERTY_348"]
            content = None
            for key in keys:
                if key in data:
                    if "TEXT" in data[key]:
                        content = list(data[key]["TEXT"].values())[0]
                        if "TYPE" in data[key]:
                            content_type = list(data[key]["TYPE"].values())[0]

                    elif "TEXT" in list(data[key].values())[0]:
                        content = list(data[key].values())[0]["TEXT"]
                        if "TYPE" in list(data[key].values())[0]:
                            content_type = list(data[key].values())[0]["TYPE"]

        if "ACTIVE_FROM" in data:
            date_publiction = data['ACTIVE_FROM']
            if data["ACTIVE_FROM"] == None:
                data["active"] = False
            # data.pop('ACTIVE_FROM')
        else:
            date_publiction = None

        if "DATE_CREATE" in data:
            date_creation = data['DATE_CREATE']
            # data.pop('DATE_CREATE')
        elif "PROPERTY_665" in data:
            date_creation = list(data['PROPERTY_665'].values())[0]
        elif "PROPERTY_666" in data:
            date_creation = list(data['PROPERTY_666'].values())[0]
        else:
            date_creation = None

        # записываем файлы в БД
        await self.search_files(data["IBLOCK_ID"], self.id, data, session)

        # article_data["indirect_data"]["files"]

        # определяем превью

        # тут, по необходимости, можно форматировать data (заменить числовой ключ на значение или что-то вроде того)

        # убрать ключи из PROPERTY:
        for key in data.keys():
            if key.startswith("PROPERTY_") and type(data[key]) == type(dict()):
                grya = []
                for key_key in data[key].keys():
                    if type(data[key][key_key]) == type(list()):
                        for scr_scr in data[key][key_key]:
                            grya.append(scr_scr)
                    else:
                        grya.append(data[key][key_key])
                data[key] = grya

        # отдельно обарботаем случай Доски почета
        if self.section_id == 14:
            # соберём совою indirect_data
            if type(data['PROPERTY_1036']) == type(list()):
                uuid = int(data['PROPERTY_1036'][0])
            else:
                uuid = int(list(data['PROPERTY_1036'].values())[0])

            if type(data['PROPERTY_1035']) == type(list()):
                year = data['PROPERTY_1035'][0]
            else:
                year = list(data['PROPERTY_1035'].values())[0]

            if type(data['PROPERTY_1037']) == type(list()):
                position = data['PROPERTY_1037'][0]
            else:
                position = list(data['PROPERTY_1037'].values())[0]

            if type(data['PROPERTY_1039']) == type(list()):
                department = data['PROPERTY_1039'][0]
            else:
                department = list(data['PROPERTY_1039'].values())[0]

            if "PROPERTY_1113" in data:
                if type(data['PROPERTY_1113']) == type(list()):
                    pre_award = data['PROPERTY_1113'][0]
                else:
                    pre_award = list(data['PROPERTY_1113'].values())[0]
                award = "Почетная грамота" if int(pre_award) == 889 else "Сотрудник года"
            else:
                award = "Сотрудник года"

            user = await User(id=uuid).search_by_id_all(session)
            if "photo_file_url" not in user or user["photo_file_url"] == None:
                photo_replace = "https://portal.emk.ru/local/templates/intranet/img/no-user-photo.png"
            else:
                photo = user["photo_file_url"]
                photo_replace = photo.replace("user_files", "compress_image/user")
            indirect_data = {
                "uuid": uuid,
                "year": year,
                "position": position,
                "department": department,
                # внедряю компрессию
                "photo_file_url": photo_replace,
                "award": award,
                "location": "Центральный офис"
            }

        # Наши люди
        elif self.section_id == 13:
            user_uuids = None
            if "PROPERTY_1235" in data:
                user_uuids = data["PROPERTY_1235"]

            indirect_data = {
                "user_uuids": user_uuids,
            }

        # отдельно обработаем случай конкурсов ЭМК
        elif self.section_id == 7:
            property_dict = {
                "CREATED_BY": "author",
                "PROPERTY_391": "sectionHref"
            }

            indirect_data = dict_to_indirect_data(data, property_dict)

        elif self.section_id == 71:
            nomination = None
            age_group = None
            # property_dict = {
            #     "PROPERTY_1071" : "nomination",
            #     "PROPERTY_1072" : "age_group",
            #     "PROPERTY_1070" : "author",
            #     "created_by" : "CREATED_BY",
            #     "PROPERTY_1074" : "representative_id",
            #     "representative_text" : "PROPERTY_1075",
            #     "PROPERTY_1073" : "likes_from_b24"
            # }

            # indirect_data = dict_to_indirect_data(data, property_dict)
            # print(data)
            if 'PROPERTY_1071' in data:
                if int(data['PROPERTY_1071'][0]) == 664:
                    nomination = 'Дети от 5 до 7 лет'
                elif int(data['PROPERTY_1071'][0]) == 1775:
                    nomination = 'Дети от 8 до 11 лет'
                elif int(data['PROPERTY_1071'][0]) == 1776:
                    nomination = 'Дети от 12 до 16 лет'
                elif int(data['PROPERTY_1071'][0]) == 2162:
                    nomination = 'Наше лето'
                elif int(data['PROPERTY_1071'][0]) == 2182:
                    nomination = 'Арматура как объект искусства'

            if 'PROPERTY_1072' in data:
                if int(data['PROPERTY_1072'][0]) == 671:
                    age_group = 'Дети от 5 до 7 лет'
                elif int(data['PROPERTY_1072'][0]) == 672:
                    age_group = 'Дети от 8 до 11 лет'
                elif int(data['PROPERTY_1072'][0]) == 673:
                    age_group = 'Дети от 12 до 16 лет'

            # indirect_data = json.dumps({
            #     "created_by" : data['CREATED_BY'],
            #     "author" : str(data['PROPERTY_1070'][0]),
            #     "nomination" : nomination,
            #     "age_group" : age_group,
            #     "representative_id" : int(data['PROPERTY_1074'][0]),
            #     "representative_text" : str(data['PROPERTY_1075'][0])
            # })

            '''!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!'''

            indirect_data = {
                "created_by": data['CREATED_BY'],
                "author": str(data['PROPERTY_1070'][0]),
                "nomination": nomination,
                "age_group": age_group,
                "representative_id": int(data['PROPERTY_1074'][0]),
                "representative_text": str(data['PROPERTY_1075'][0]) if 'PROPERTY_1075' in data.keys() else None,
                "likes_from_b24": data['PROPERTY_1073']
            }

        # отдельно обарботаем случай Блогов
        elif self.section_id == 15:
            # собираем из двух статей одну
            uuid = None
            photo = None
            author = None
            if "PROPERTY_444" in data:
                if type(data['PROPERTY_444']) == type(list()):
                    uuid = int(data['PROPERTY_444'][0])
                else:
                    uuid = int(list(data['PROPERTY_444'].values())[0])

                # отдельно вытащить превьюшки людей
                user = await User(id=uuid).search_by_id_all(session)
                
                photo = user["photo_file_url"]
                # photo = photo.replace("user_files", "compress_image/user")

                # ФИО
                fio = user['last_name'] + " " + user['name'] + " " + user['second_name']

                # взять должность
                if "work_position" in user["indirect_data"]:
                    position = user["indirect_data"]['work_position']
                else:
                    position = ""

                author = fio + ";" + position

            company = None
            if "PROPERTY_1022" in data and take_value(data["PROPERTY_1022"]) == "6180":
                company = 10834  # "АО «НПО «Регулятор»"
            elif "PROPERTY_1022" in data and take_value(data["PROPERTY_1022"]) == "6178":
                company = 10815  # "АО «САЗ»"

            if "PROPERTY_453" in data and take_value(data["PROPERTY_453"]) == "335":
                data["active"] = True
            else:
                data["active"] = False

            if "PROPERTY_446" in data and take_value(data["PROPERTY_446"]) == "333":
                data["active"] = True
            else:
                data["active"] = False

            link = None
            if "PROPERTY_1247" in data:
                link = take_value(data["PROPERTY_1247"])

            YouTube = None
            if "PROPERTY_1222" in data:
                YouTube = take_value(data["PROPERTY_1222"])

            # отдельно обрабатываем файлы
            if "PROPERTY_1239" in data:
                content = take_value(data["PROPERTY_1239"])
            if content is not None:
                # хватаю url
                matches = re.findall(r'src="([^"]*)"', content)
                for url in matches:
                    # качаю файл новым методом
                    if url != "https://portal.emk.ru/bitrix/tools/disk/uf.php?attachedId=128481&auth%5Baplogin%5D=1&auth%5Bap%5D=j6122m0ystded5ag&action=show&ncc=1":
                        url_b24 = 'https://portal.emk.ru' + url
                        new_url = await File().upload_by_URL(url=url_b24, art_id=self.id,
                                                             session=session)  # СЮДА АСИНХРОННОСТЬ
                        # print(url_b24, "-->", new_url)
                        # заменяю url на новый
                        # content = re.sub(r'src="([^"]*)"', f'src="{new_url}"', content)

                        content = content.replace(url, new_url)

            indirect_data = {
                "TITLE": data["TITLE"],
                "author_uuid": uuid,
                "author": author,
                "company": company,
                "link": link,
                "youtube_link": YouTube,
                "photo_file_url": photo,
            }

            # файлы для Интранета ???сработает??? - да

            keys = [
                "PROPERTY_1023",  # фото превью
                "PROPERTY_1222",  # ссылка на youtube
                "PROPERTY_455",
                "PROPERTY_1020",
            ]
            for key in keys:
                if key in data:
                    indirect_data[key] = data[key]

        # видеоинтервью
        elif self.section_id == 16:
            author = None
            if "PROPERTY_1026" in data:
                author = data["PROPERTY_1026"]

            indirect_data = {"author": author}


        # отдельно забираю сортировку для Памятки Новому Сотруднику
        elif self.section_id == 18:
            sort = None
            if "PROPERTY_475" in data:
                sort = take_value(data["PROPERTY_475"])
            indirect_data = {"sort": sort}

        # Референсы и опыт поставок
        elif self.section_id == 25:

            industryId = None
            if "PROPERTY_681" in data:
                industryId = take_value(data["PROPERTY_681"])

            industry = None
            values_dict = {
                None: "Прочие",
                "8308": "Прочие",
                "8307": "Энергетика",
                "8306": "Химия",
                "8305": "Нефтегаз"
            }
            industry = values_dict[industryId]

            enterpriseId = None
            if "PROPERTY_680" in data:
                enterpriseId = take_value(data["PROPERTY_680"])

            enterprise = None
            values_dict = {
                None: "Ошибка",
                "6185": "ООО «Пульсатор»",
                "6184": "ООО «Техно-Сфера»",
                "6183": "ООО «АРМАТОМ»",
                "6182": "АО «Тулаэлектропривод»",
                "6181": "ООО «ТехПромАрма»",
                "6180": "АО «НПО Регулятор»",
                "6179": "ЗАО «Курганспецарматура»",
                "6178": "ЗАО «Саратовский арматурный завод»"
            }
            enterprise = values_dict[enterpriseId]

            indirect_data = {
                "industry": industry,
                "industryId": industryId,
                "enterprise": enterprise,
                "enterpriseId": enterpriseId
            }

        # Актуальные новости и Корпоративные события
        elif self.section_id == 31 or self.section_id == 51:
            indirect_data = {}
            author = None
            if "PROPERTY_294" in data:
                author = data["PROPERTY_294"]
            else:
                pass

            if "PROPERTY_1116" in data and self.section_id == 31:
                tags = []
                for value in data['PROPERTY_1116']:
                    existing_tag = await Tag(id=int(value)).get_tag_by_id(session)
                    if existing_tag:
                        tags.append(int(value))
                indirect_data['tags'] = tags

            indirect_data["author"] = author



        # Благотворительные проекты
        elif self.section_id == 55:
            property_dict = {
                "PROPERTY_435": "organizer",
                "PROPERTY_347": "phone_number",
                "PROPERTY_344": "theme"
            }

            indirect_data = dict_to_indirect_data(data, property_dict)

        # Учебный центр (Литература)
        elif self.section_id == 175:
            property_dict = {
                "PROPERTY_489": "subsection_id",
                "PROPERTY_488": "author"
            }

            indirect_data = dict_to_indirect_data(data, property_dict)

            subsection_id = indirect_data["subsection_id"]
            values_dict = {
                None: "Нет данных",
                "339": "Техническая литература",
                "340": "Обучающие материалы",
                "1020": "Диджитал и IT",
                "1021": "Психология и развитие",
                "1761": "Обучающие материалы: продажи B2B",
                "1762": "Обучающие материалы: Эффективные переговоры",
                "1763": "Обучающие материалы: Профессиональное планирование для регулярного менеджмента",
            }
            indirect_data["subsection"] = values_dict[subsection_id]

        # Учебный центр (Тренинги)
        elif self.section_id == 172:
            if "PROPERTY_371" in data:
                content = data["PROPERTY_371"][0]["TEXT"]
                content_type = data["PROPERTY_371"][0]["TYPE"]

            property_dict = {
                "PROPERTY_369": "event_date",
                "PROPERTY_437": "author",
                "PROPERTY_432": "participants"
            }

            indirect_data = dict_to_indirect_data(data, property_dict)
            participants = []
            if "participants" in indirect_data:
                for user_uuid in indirect_data["participants"]:
                    user = await User(id=user_uuid).search_by_id_all(session)
                    if user is not None:
                        last_name = user['last_name']
                        name = user['name']
                        second_name = user['second_name']

                        fio = f"{last_name} {name} {second_name}"
                        photo = user["photo_file_url"]
                        work_position = user["indirect_data"]["work_position"]

                        participants.append({
                            "fio": fio,
                            "photo_file_url": photo,
                            "position": work_position
                        })

            reviews_props = data["reviews"]
            reviews = []
            if reviews_props != []:
                for feedback_props in reviews_props:
                    text = ""
                    if "PROPERTY_486" in feedback_props:
                        text = list(feedback_props["PROPERTY_486"].values())[0]["TEXT"]

                    name = "",
                    if "NAME" in feedback_props:
                        name = feedback_props["NAME"]

                    stars = "",
                    if "PROPERTY_501" in feedback_props:
                        stars = list(feedback_props["PROPERTY_501"].values())[0]
                        print(feedback_props["PROPERTY_501"], stars)

                    feedback = {
                        "reviewer": name,
                        "text": text,
                        "stars": stars,
                    }
                    reviews.append(feedback)

            indirect_data["reviews"] = reviews
            indirect_data["participants"] = participants

        # Новости организационного развития
        elif self.section_id == 32:

            indirect_data = {"users": [], "active_main_page": False}
            if preview is not None and preview != "":
                users_data = extract_user_data(preview)
                for user_data in users_data:
                    # нати данные пользователя
                    usr_id = user_data['id']
                    usr = await User(id=usr_id).search_by_id_all(session)

                    # ФИО
                    fio = usr['last_name'] + " " + usr['name'] + " " + usr['second_name']
                    # фото
                    photo_file_url = usr["photo_file_url"]
                    # взять должность
                    position = user_data['text']
                    usr = {
                        "id": usr_id,
                        "fio": fio,
                        "photo_file_url": photo_file_url,
                        "position": position
                    }
                    indirect_data["users"].append(usr)

        # Корпоративная газета ЭМК
        elif self.section_id == 34:
            img_url = await File().upload_by_URL(url=data["image"], art_id=self.id, is_preview=True, session=session)
            file_url = await File().upload_by_URL(url=data["file"], art_id=self.id, session=session)
            indirect_data = {
                "year": data["year"],
                "photo_file_url": img_url,
                "pdf": file_url,
            }

        # Гид по предприятиям
        elif self.section_id == 41:

            report = data["reports"]
            tour = data["tours"]

            reports = []
            tours = []

            print(reports)
            if report != []:
                for rep in report:
                    act = True
                    if rep["BP_PUBLISHED"] != "Y":
                        act = False

                    photo_file_url = None
                    if "PROPERTY_669" in rep:
                        photo = take_value(rep["PROPERTY_669"])
                        print(photo)
                        # скачать и вытащить ссылку
                        files = [photo]
                        art_id = self.id
                        inf_id = "98"
                        is_preview = False

                        file_data = await File(b24_id=photo).upload_inf_art(art_id=art_id, is_preview=is_preview,
                                                                            need_all_method=True, inf_id=inf_id,
                                                                            session=session)
                        print(file_data)

                        if file_data is None:
                            photo_file_url = None

                        else:
                            url = file_data["file_url"]
                            photo_file_url = f"{DOMAIN}{url}"

                    rp = {
                        "rep_id": rep["ID"],
                        "name": rep["NAME"],
                        "active": act,
                        "date": take_value(rep["PROPERTY_667"]),
                        "photo_file_url": photo_file_url,
                        "link": take_value(rep["PROPERTY_670"])  # !!!!!!!!!!!!!! сслыка на youtube
                    }

                    reports.append(rp)

            if tour != []:
                for tr in tour:
                    act = True
                    if tr["BP_PUBLISHED"] != "Y":
                        act = False
                    
                    if tr["ID"] == "7442" or tr["ID"] == "7441":
                        act = False
                        photo_file_url = None

                    if "PROPERTY_498" in tr:
                        photo = take_value(tr["PROPERTY_498"])
                        # скачать и вытащить ссылку
                        art_id = self.id
                        inf_id = "84"
                        is_preview = False

                        
                        # ЕСЛИ ФАЙЛЫ БЫЛИ СКАЧЕННЫ РАНЕЕЕ ТО ОН ПЕРЕЗАПИШЕТ НА NULL
                        if act:
                            file_data = await File(b24_id=photo).upload_inf_art(art_id=art_id, is_preview=is_preview,
                                                                                need_all_method=True, inf_id=inf_id,
                                                                                session=session)
                        
                            if file_data:
                                url = file_data["file_url"]
                                photo_file_url = f"{DOMAIN}{url}"

                        # if tr["ID"] == "7442" or tr["ID"] == "7596":
                        #     print(photo_file_url, photo)

                    t = {
                        "tourId": tr["ID"],
                        "factory_id": self.id,
                        "name": tr["NAME"],
                        "active": act,
                        "3D_files_path": take_value(tr["PROPERTY_497"]),
                        "photo_file_url": photo_file_url
                    }

                    tours.append(t)

            indirect_data = {
                "PROPERTY_463": data["PROPERTY_463"],
                "reports": reports,
                "tours": tours
            }

        # Галерея фото и видео
        elif self.section_id == 42 or self.section_id == 52:
            indirect_data = dict()

        # Афиша
        elif self.section_id == 53:
            property_dict = {
                "PROPERTY_375": "date_from",
                "PROPERTY_438": "date_to"
            }

            indirect_data = dict_to_indirect_data(data, property_dict)

        # Вакансии (приведи друга)
        elif self.section_id == 111:
            property_dict = {
                "PROPERTY_5094": "link"
            }

            indirect_data = dict_to_indirect_data(data, property_dict)

        # Предложения партнеров
        elif self.section_id == 54:

            indirect_data = dict()

        # Видеорепортажи
        elif self.section_id == 33:
            if "PROPERTY_1116" in data:
                indirect_data = data
                tags = []
                for value in data['PROPERTY_1116']:
                    existing_tag = await Tag(id=int(value)).get_tag_by_id(session)
                    if existing_tag:
                        tags.append(int(value))
                indirect_data['tags'] = tags

        else:
            indirect_data = json.dumps(data)

        article_data = {
            "id": self.id,
            "section_id": self.section_id,
            "name": data['NAME'],
            "preview_text": preview,
            "content_text": content,
            "date_publiction": make_date_valid(date_publiction),
            "date_creation": make_date_valid(date_creation),
            "indirect_data": indirect_data
        }

        if "active" in data:
            article_data['active'] = data['active']

        if content_type is not None:
            article_data['content_type'] = content_type

        return article_data

    async def search_files(self, inf_id, art_id, data, session):
        files_propertys = [
            "PREVIEW_PICTURE",
            "DETAIL_PICTURE",

            "PROPERTY_372",
            "PROPERTY_373",

            "PROPERTY_337",
            "PROPERTY_338",

            "PROPERTY_342",
            "PROPERTY_343",

            # Блоги
            "PROPERTY_1023",
            "PROPERTY_1222",  # ссылка на youtube
            "PROPERTY_1203",  # ссылка на youtube
            "PROPERTY_455",
            "PROPERTY_1020",
            "PROPERTY_1246",  # QR-код Земской

            # Референсы
            "PROPERTY_678",
            # "PROPERTY_679",

            "PROPERTY_476",

            # Актуальные новости и Корпоративные события
            "PROPERTY_491",
            "PROPERTY_664",  # ссылка на youtube

            # "PROPERTY_670", #!!! сслыка на ютуб !!!
            "PROPERTY_669",

            # Гид по предприятиям
            "PROPERTY_463",

            "PROPERTY_498",

            "PROPERTY_289",
            # "PROPERTY_296",

            "PROPERTY_399",

            "PROPERTY_400",
            # "PROPERTY_402",
            "PROPERTY_407",

            "PROPERTY_409",  # !!! сслыка на ютуб !!!

            "PROPERTY_476",
            "PROPERTY_1025",
            "PROPERTY_356",

            # вложения
            "PROPERTY_478",
            "PROPERTY_491",
            "PROPERTY_366",

            # превьюшка конкурсов
            "PROPERTY_389",
        ]

        preview_file = [
            "PROPERTY_399",
            "PROPERTY_407",
            "PROPERTY_372",
            "PROPERTY_337",
            "PROPERTY_342",
            "PROPERTY_476",
            "PROPERTY_669",
            "PROPERTY_463",
            "PROPERTY_498",
            "PREVIEW_PICTURE",
            "B24_PREVIEW_FILES",
            "PROPERTY_356",
            "PROPERTY_389",
        ]

        link_prop = [
            "PROPERTY_664",
            "PROPERTY_1222",
            "PROPERTY_1203",
            "PROPERTY_670",
            "PROPERTY_409"
        ]

        default_flase = [
            "PROPERTY_289",
            "PROPERTY_400",
            "PROPERTY_373",
            "PROPERTY_678",
            "PROPERTY_366"
        ]

        files_data = []
        # прохожу по всем проперти статьи
        for file_property in files_propertys:
            # если это файловый проперти
            if file_property in data:
                # если это ссылка
                if file_property in ["PROPERTY_664", "PROPERTY_1222", "PROPERTY_1203", "PROPERTY_670", "PROPERTY_409"]:
                    # !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!  проверка есть ли в битре такая ссылка или нет
                    link = take_value(data[file_property])
                    f_res = await File(b24_id=f"link_{art_id}").add_link(link=link, art_id=art_id, session=session)
                    files_data.append(f_res)

                # если это файл превью
                elif file_property in preview_file:

                    preview_images = []
                    if type(data[file_property]) == type(dict()):
                        for file_id in data[file_property].values():
                            if type(file_id) == type(str()):
                                preview_images.append(file_id)
                            elif type(file_id) == type(list()):
                                for f_id in file_id:
                                    preview_images.append(f_id)
                    elif type(data[file_property]) == type(list()):
                        for dct in data[file_property]:
                            for file_id in dct.values():
                                if type(file_id) == type(str()):
                                    preview_images.append(file_id)
                                elif type(file_id) == type(list()):
                                    for f_id in file_id:
                                        preview_images.append(f_id)
                    elif type(data[file_property]) == type(str()):
                        preview_images.append(data[file_property])

                    files_to_add = await File().need_update_file(art_id=art_id, files_id=preview_images,
                                                                 session=session)

                    if files_to_add != []:
                        for f_id in files_to_add:

                            try:
                                LogsMaker().info_message(
                                    f" Качаю файл превью {f_id} статьи {art_id} инфоблока {inf_id}, использование метода Матренина - ДА")
                                file_data = await File(b24_id=f_id).upload_inf_art(art_id=art_id, is_preview=True,
                                                                                   need_all_method=True, inf_id=inf_id,
                                                                                   session=session)
                                files_data.append(file_data)
                            except:
                                LogsMaker().info_message(
                                    f" Качаю файл превью {f_id} статьи {art_id} инфоблока {inf_id}, использование метода Матренина - НЕТ")
                                file_data = await File(b24_id=f_id).upload_inf_art(art_id=art_id, is_preview=True,
                                                                                   need_all_method=False, inf_id=inf_id,
                                                                                   session=session)
                                files_data.append(file_data)

                                # остальные файлы
                else:
                    need_all_method = True
                    if file_property in ["PROPERTY_289", "PROPERTY_400", "PROPERTY_373", "PROPERTY_678",
                                         "PROPERTY_366"]:
                        need_all_method = False

                    files = []
                    if type(data[file_property]) == type(dict()):
                        for file_id in data[file_property].values():
                            if type(file_id) == type(str()):
                                files.append(file_id)
                            elif type(file_id) == type(list()):
                                for f_id in file_id:
                                    files.append(f_id)
                    elif type(data[file_property]) == type(list()):
                        for dct in data[file_property]:
                            for file_id in dct.values():
                                if type(file_id) == type(str()):
                                    files.append(file_id)
                                elif type(file_id) == type(list()):
                                    for f_id in file_id:
                                        files.append(f_id)
                    elif type(data[file_property]) == type(str()):
                        files.append(data[file_property])

                    files_to_add = await File().need_update_file(art_id=art_id, files_id=files, session=session)

                    if files_to_add != []:
                        for f_id in files_to_add:
                            msg = f" Качаю файл {f_id} статьи {art_id} инфоблока {inf_id}, использование метода Матренина - {need_all_method}"
                            LogsMaker().info_message(msg)
                            try:
                                file_data = await File(b24_id=f_id).upload_inf_art(art_id=art_id, is_preview=False,
                                                                                   need_all_method=need_all_method,
                                                                                   inf_id=inf_id, session=session)
                                files_data.append(file_data)
                            except:
                                LogsMaker().warning_message(
                                    f" Не получилось по хорошему скачать файл {f_id} статьи {art_id} инфоблока {inf_id}, метода Матренина по умолчанию - {need_all_method}")
                                file_data = await File(b24_id=f_id).upload_inf_art(art_id=art_id, is_preview=False,
                                                                                   need_all_method=not need_all_method,
                                                                                   inf_id=inf_id, session=session)
                                files_data.append(file_data)

        return files_data

    '''
    async def old_search_files(self, inf_id, art_id, data):

        files_propertys = [
            "PREVIEW_PICTURE",
            "DETAIL_PICTURE",

            "PROPERTY_372",
            "PROPERTY_373",

            "PROPERTY_337",
            "PROPERTY_338",

            "PROPERTY_342",
            "PROPERTY_343",

            #Блоги
            "PROPERTY_1023", 
            "PROPERTY_1222", #ссылка на youtube
            "PROPERTY_1203", #ссылка на youtube
            "PROPERTY_455",
            "PROPERTY_1020",
            "PROPERTY_1246", #QR-код Земской

            #Референсы
            "PROPERTY_678",
            #"PROPERTY_679",

            "PROPERTY_476",

            # Актуальные новости и Корпоративные события
            "PROPERTY_491",
            "PROPERTY_664", #ссылка на youtube

            #"PROPERTY_670", #!!! сслыка на ютуб !!!
            "PROPERTY_669",

            #Гид по предприятиям
            "PROPERTY_463",

            "PROPERTY_498",

            "PROPERTY_289",
            # "PROPERTY_296",

            "PROPERTY_399",

            "PROPERTY_400",
            #"PROPERTY_402",
            "PROPERTY_407",

            "PROPERTY_409", #!!! сслыка на ютуб !!!

            "PROPERTY_476",
            "PROPERTY_1025",
            "PROPERTY_356",

            #вложения
            "PROPERTY_478",
            "PROPERTY_491",
            "PROPERTY_366",
        ]

        preview_file = [
            "PROPERTY_399",
            "PROPERTY_407",
            "PROPERTY_372",
            "PROPERTY_337",
            "PROPERTY_342",
            "PROPERTY_476",
            "PROPERTY_669",
            "PROPERTY_463",
            "PROPERTY_498",
            "PREVIEW_PICTURE",
            "B24_PREVIEW_FILES",
            "PROPERTY_356",
        ]



        # находим файлы статьи
        files = []
        preview_images = []
        files_data = []
        #собираем данные о файлах
        for file_property in files_propertys:
            need_all_method = True
            if file_property in data:
                # if art_id == 12221:
                #     print(data, art_id)

                #ссылки 
                if file_property in ["PROPERTY_664", "PROPERTY_1222", "PROPERTY_1203", "PROPERTY_670", "PROPERTY_409"]:
                    link = take_value(data[file_property])
                    File(b24_id=f"link_{art_id}").add_link(link, art_id)

                #обрабатываются дефолтным методом битры
                if file_property in ["PROPERTY_289", "PROPERTY_400", "PROPERTY_373", "PROPERTY_678", "PROPERTY_366"]:
                    need_all_method = False
                elif file_property in ["PROPERTY_491"]:
                    need_all_method = True
                try:
                    # выцепить id файла
                    # "PREVIEW_PICTURE" не обрабатывается, тип - строка
                    # "DETAIL_PICTURE" тоже не обработается если строка
                    if type(data[file_property]) == type(dict()):
                        for file_id in data[file_property].values():
                            if type(file_id) == type(str()):
                                files.append(file_id)
                                if file_property in preview_file:
                                        preview_images.append(file_id)
                            elif type(file_id) == type(list()):
                                for f_id in file_id:
                                    files.append(f_id)
                                    if file_property in preview_file:
                                        preview_images.append(f_id)
                    elif type(data[file_property]) == type(list()):
                        for dct in data[file_property]:
                            for file_id in dct.values():
                                if type(file_id) == type(str()):
                                    files.append(file_id)
                                    if file_property in preview_file:
                                            preview_images.append(file_id)
                                elif type(file_id) == type(list()):
                                    for f_id in file_id:
                                        files.append(f_id)
                                        if file_property in preview_file:
                                            preview_images.append(f_id)
                    elif type(data[file_property]) == type(str()):
                        files.append( data[file_property] )

                        if file_property in preview_file:
                            preview_images.append(data[file_property])
                    else:
                        LogsMaker().warning_message("Некорректные данные в поле ", file_property, f"Данные: {type(data[file_property])}", f"Ищи в {inf_id}, {art_id}")
                        # print("Некорректные данные в поле ", file_property, f"Данные: {type(data[file_property])}", f"Ищи в {inf_id}, {art_id}")

                except Exception as e:
                    return LogsMaker().error_message(e)
                    # print("Ошибка обработки в инфоблоке", sec_inf[i], "в поле", file_property)

            if files != []:

                files_data = []
                #проеверяем, нужно ли обновить файлы?
                # if art_id == 12221:
                #     print(f'{files} проверяет на обновлениеб {preview_images} - сработали ли?')

                files_to_add = File().need_update_file(art_id, files)

                if files_to_add != []:
                    for f_id in files:
                        print(f"Качаю файл {f_id} статьи {art_id} инфоблока {inf_id}, использование метода Матренина - {need_all_method}")
                        try:
                            is_preview = f_id in preview_images
                            file_data = File(b24_id=f_id).upload_inf_art(art_id, is_preview, need_all_method, inf_id)
                            #sprint(f'{f_id} файл добавлен в монго', art_id, inf_id)
                            files_data.append(file_data)
                        except:
                            LogsMaker().warning_message(f"Не получилось по хорошему скачать файл {f_id} статьи {art_id} инфоблока {inf_id}, метода Матренина по умолчанию - {need_all_method}")
                            is_preview = f_id in preview_images
                            file_data = File(b24_id=f_id).upload_inf_art(art_id, is_preview, True, inf_id)
                            # sprint(f'{f_id} файл добавлен в монго', art_id, inf_id)
                            files_data.append(file_data)

            return files_data
    '''

    async def add(self, article_data, session):
        return await ArticleModel().add_article(article_data=await self.make_valid_article(article_data, session),
                                                session=session)

    async def set_new(self, article_data, session):
        return await ArticleModel().add_article(article_data=article_data, session=session)

    async def uplod(self, session):
        '''
        ! Не повредить имеющиеся записи и структуру
        ! Выгрузка файлов из инфоблоков
        ✔️ - готов и отлажен
        ☑️ - готов и тестируется
        ♻️ - в разработке сейчас
        ❌ - ожидает работы
        '''

        '''
        ! Сопоставить section_id из Интранета и IBLOCK_ID из B24
        '''

        await self.upload_uniquely(session)
        # await self.upload_with_parameter(session)
        await self.upload_many_to_many(session)
        # await self.upload_services(session)  # загрузили все без проблем

        # Дамп данных в эластик
        await self.dump_articles_data_es(session=session)

        await self.upload_likes(session)
        await self.upload_views(session)

        return {'status': True}

    async def upload_uniquely(self, session):
        '''однозначно'''
        sec_inf = {
            # 13: "149",  # Наши люди ✔️ DONE
            # 14: "123",  # Доска почёта ✔️ DONE
            # 16: "122",  # Видеоитервью ✔️ DONE

            # 32: "132",  # Новости организационного развития ✔️  DONE
            53: "62",  # Афиша ✔️ DONE
            # 54: "55",  # Предложения партнеров ✔️ DONE
            # 55: "56",  # Благотворительные проекты ✔️  DONE

            # 25: "100",  # Референсы и опыт поставок ✔️DONE
            # 175: "60",
            # Учебный центр (Литература) ✔️ DONE (но не скачались по вине битры 23038, 23041, 23044, 23134, 23137, 23141, 23149, 23151)
            7: "66",  # Конкурсы (Главная) ✔️
            71: "128",  # Конкурсы (Непосредственно)
        }

        # проходимся по инфоблокам
        for i in self.logg.progress(sec_inf, f"Загрузка данных инфоблоков {sec_inf.values} "):

            # запрос в B24
            self.section_id = sec_inf[i]
            infs = await self.get_inf()

            # инфоблок не пустой
            if infs != []:
                for inf in infs:
                    artDB = ArticleModel(id=inf["ID"], section_id=i)
                    self.section_id = i
                    if await artDB.need_add(session=session):
                        self.logg.info_message(f'Добавил статью, {inf["ID"]}')
                        await self.add(inf, session)
                    elif await artDB.update(article_data=await self.make_valid_article(inf, session), session=session):
                        # проверить апдейт файлов
                        pass

    async def upload_with_parameter(self, session):
        '''с параметрами'''
        # один section_id - несколько IBLOCK_ID
        sec_inf = {
            15: ["75", "77"],  # Блоги ✔️
            18: ["81", "82"],  # Памятка ✔️
            41: ["98", "78", "84"],  # Гид по предприятиям ✔️ сделать сервис
            172: ["61", "83"]  # Учебный центр (Проведённые тренинги) ✔️
        }

        # # # Учебный центр (Проведённые тренинги)
        # # self.section_id = "61"
        # sec_inf_title = await self.get_inf()
        # for title_inf in self.logg.progress(sec_inf_title, "Загрузка данных инфоблоков 61, 83 "):
        #     title_id = title_inf["ID"]
        #     title_data = title_inf

        #     data = dict()

        #     # добавить все данные статьи
        #     for key in title_data:
        #         data[key] = title_data[key]

        #     data["ID"] = title_data["ID"]
        #     data["TITLE"] = title_data["NAME"]
        #     # print(data["ID"], data)
        #     data["reviews"] = []

        #     # пройти по инфоблоку тренингов
        #     self.section_id = "83"
        #     sec_inf_data = await self.get_inf()
        #     for data_inf in sec_inf_data:
        #         # если эта статья принадлежит иинфоблоку
        #         if "PROPERTY_484" in data_inf and take_value(data_inf["PROPERTY_484"]) == title_id:
        #             # добавить отзывы
        #             data["reviews"].append(data_inf)

        #     # загрузить данные в таблицу
        #     data["section_id"] = 172
        #     self.section_id = 172
        #     if int(data["ID"]) == 10855:
        #         print(data, 'тут пустышки')
        #     artDB = ArticleModel(id=data["ID"], section_id=self.section_id)
        #     if await artDB.need_add(session=session):
        #         await self.add(data, session)
        #     elif await artDB.update(await self.make_valid_article(data, session), session):
        #         pass

        # # # Блоги
        # # # пройти по инфоблоку заголовков
        # self.section_id = "75"
        # sec_inf_title = await self.get_inf()
        # for title_inf in self.logg.progress(sec_inf_title, "Загрузка данных инфоблоков 75, 77 "):
        #     title_id = title_inf["ID"]
        #     title_data = title_inf

        #     # пройти по инфоблоку статей блогов
        #     self.section_id = "77"
        #     sec_inf_data = await self.get_inf()
        #     for data_inf in sec_inf_data:
        #         data_title_id = list(data_inf["PROPERTY_1008"].values())[0]
        #         # если эта статья принадлежит иинфоблоку
        #         if data_title_id == title_id:
        #             data = dict()

        #             # добавить все данные заголовка
        #             for key in title_data:
        #                 data[key] = title_data[key]
        #             # добавить все данные статьи
        #             for key in data_inf:
        #                 data[key] = data_inf[key]

        #             data["ID"] = data_inf["ID"]
        #             data["section_id"] = 15  # Блоги
        #             self.section_id = 15
        #             data["TITLE"] = title_inf["NAME"]
        #             if int(data["ID"]) == 10855:
        #                 print(data, 'тут пустышки')
        #             # загрузить данные в таблицу
        #             artDB = ArticleModel(id=int(data["ID"]), section_id=self.section_id)
        #             if await artDB.need_add(session=session):
        #                 await self.add(data, session)
        #             elif await artDB.update(await self.make_valid_article(data, session), session):
        #                 pass

        # # # # Памятка
        # # # # пройти по инфоблоку заголовков
        # self.section_id = "82"
        # sec_inf_title = await self.get_inf()
        # for title_inf in self.logg.progress(sec_inf_title, "Загрузка данных инфоблоков 82, 81 "):
        #     title_id = title_inf["ID"]
        #     title_data = title_inf

        #     # пройти по инфоблоку статей блогов
        #     self.section_id = "81"
        #     sec_inf_data = await self.get_inf()
        #     for data_inf in sec_inf_data:
        #         if "PROPERTY_480" in data_inf:
        #             data_title_id = list(data_inf["PROPERTY_480"].values())[0]
        #         else:
        #             self.logg.info_message(f'##################, {data_inf["ID"]}')

        #         # если эта статья принадлежит инфоблоку
        #         if data_title_id == title_id:
        #             data = dict()

        #             # добавить все данные заголовка
        #             for key in title_data:
        #                 data[key] = title_data[key]
        #             # добавить все данные статьи
        #             for key in data_inf:
        #                 data[key] = data_inf[key]

        #             data["ID"] = data_inf["ID"]
        #             data["section_id"] = 18  # Памятка
        #             self.section_id = 18
        #             data["TITLE"] = title_inf["NAME"]
        #             if int(data["ID"]) == 6180:
        #                 print(data, 'тут начинается')
        #             # загрузить данные в таблицу
        #             artDB = ArticleModel(id=int(data["ID"]), section_id=self.section_id)
        #             if await artDB.need_add(session=session):
        #                 await self.add(data, session)
        #             elif await artDB.update(await self.make_valid_article(data, session), session):
        #                 pass

        # Гид по предприятиям
        # пройти по инфоблоку заголовков
        # ПРОВЕРИТЬ НЕ СКАЧАНЫ ЛИ УЖЕ ФАЙЛЫ ПРЕВЬЮ ИНАЧЕ ПЕРЕЗАПИШЕТ НА NONE
        # self.section_id = "78"
        # sec_inf_title = await self.get_inf()
        # for title_inf in self.logg.progress(sec_inf_title, "Загрузка данных инфоблоков 78, 98 и 84"):
        #     art_id = title_inf["ID"]
        #     data = title_inf
        #     data["reports"] = []
        #     data["tours"] = []

        #     # пройти по инфоблоку репортажей
        #     self.section_id = "98"
        #     sec_inf_data = await self.get_inf()
        #     for data_inf in sec_inf_data:
        #         # if "PROPERTY_671" in data_inf:
        #         data_title_id = list(data_inf["PROPERTY_671"].values())[0]
        #         # если эта статья принадлежит иинфоблоку

        #         if data_title_id == art_id:
        #             dt = dict()

        #             # добавить все данные статьи
        #             for key in data_inf:
        #                 dt[key] = data_inf[key]

        #             dt["ID"] = data_inf["ID"]
        #             dt["TITLE"] = title_inf["NAME"]

        #             data["reports"].append(dt)

        #     # пройти по инфоблоку репортажей
        #     self.section_id = "84"
        #     sec_inf_data = await self.get_inf()
        #     for data_inf in sec_inf_data:
        #         # if "PROPERTY_671" in data_inf:
        #         data_title_id = list(data_inf["PROPERTY_496"].values())[0]
        #         # если эта статья принадлежит иинфоблоку

        #         if data_title_id == art_id:
        #             dt = dict()

        #             for key in data_inf:
        #                 dt[key] = data_inf[key]

        #             dt["ID"] = data_inf["ID"]
        #             dt["TITLE"] = title_inf["NAME"]

        #             data["tours"].append(dt)

        #     data["section_id"] = 41  # Гид по предприятиям
        #     self.section_id = 41
        #     # загрузить данные в таблицу
        #     if int(data["ID"]) == 10855:
        #         print(data, 'тут пустышки')
        #     artDB = ArticleModel(id=int(data["ID"]), section_id=self.section_id)
        #     if await artDB.need_add(session=session):
        #         await self.add(data, session)
        #     elif await artDB.update(await self.make_valid_article(data, session), session):
        #         pass


    async def upload_many_to_many(self, session):
        await self.upload_current_news(session)
        await self.upload_corporate_events(session)

    async def upload_current_news(self, session):

        # несколько section_id - один IBLOCK_ID
        sec_inf = {
            31: "50",  # Актуальные новости ✔️
            51: "50"  # Корпоративные события ✔️
        }

        # пройти по инфоблоку
        self.section_id = "50"
        art_inf = await self.get_inf()
        for art in self.logg.progress(art_inf,
                                      "Загрузка данных разделов \"Актуальные новости\", \"Корпоративные события\" и \"Видеорепортажи\" "):
            if art["ID"] == '13486':
                self.logg.warning_message(f'{art["ID"]} новость которая проникает не туда')
                # print(art, ' новость')
            else:
                pass
            art_id = art["ID"]
            if "PROPERTY_1066" in art:
                pre_section_id = list(art["PROPERTY_1066"].values())[0]

                if pre_section_id == "661":
                    if "PROPERTY_5044" in art and list(art["PROPERTY_5044"].values())[0] == "1":
                        art["section_id"] = 33  # Видеорепортажи
                        self.section_id = 33

                        # artDB = ArticleModel(id=art["ID"], section_id=self.section_id)
                        # if await artDB.need_add(session=session):
                        #     await self.add(art, session)
                        # elif await artDB.update( await self.make_valid_article(art, session), session):
                        #     pass

                    else:
                        art["section_id"] = 31  # Актуальные новости
                        self.section_id = 31
                elif pre_section_id == "663":
                    art["section_id"] = 51  # Корпоративные события
                    self.section_id = 51

                artDB = ArticleModel(id=art["ID"], section_id=self.section_id)
                if await artDB.need_add(session=session):
                    await self.add(art, session)
                elif await artDB.update(await self.make_valid_article(art, session), session):
                    pass
            else:
                # че делать с уже не актуальными новостями?
                self.section_id = 6
                artDB = ArticleModel(id=art["ID"], section_id=self.section_id)
                if await artDB.need_add(session=session):
                    await self.add(art, session)
                    self.logg.warning_message(f'Статья - Name:{art["NAME"]}, id:{art["ID"]} уже не актуальна')
                    # print("Статья", art["NAME"], art["ID"], "уже не актуальна")
                elif await artDB.update(await self.make_valid_article(art, session), session):
                    # сюда надо что-то дописать
                    pass

    async def upload_corporate_events(self, session):
        # несколько section_id - несколько IBLOCK_ID
        sec_inf = {
            42: ["68", "69"],  # Официальные события ✔️
            52: ["68", "69"]  # Корпоративная жизнь в фото ✔️
        }

        # Фотогалерея
        self.section_id = "68"
        art_inf = await self.get_inf()
        for art in self.logg.progress(art_inf,
                                      "Загрузка данных разделов \"Официальные события\" и \"Корпоративная жизнь в фото\" "):
            art_id = art["ID"]

            if "PROPERTY_403" in art:
                pre_section_id = list(art["PROPERTY_403"].values())[0]

                if pre_section_id == "322":
                    art["section_id"] = 42  # Официальные события
                    self.section_id = 42
                elif pre_section_id == "323":
                    art["section_id"] = 52  # Корпоративная жизнь в фото
                    self.section_id = 52

                artDB = ArticleModel(id=art["ID"], section_id=self.section_id)
                if await artDB.need_add(session=session):
                    await self.add(art, session)
                elif await artDB.update(await self.make_valid_article(art, session), session):
                    pass

            else:
                self.section_id = 6
                artDB = ArticleModel(id=art["ID"], section_id=self.section_id)
                if await artDB.need_add(session=session):
                    await self.add(art, session)
                    # че делапть с уже не актуальными новостями?
                    print("Запись в фотогалерею", art["NAME"], art["ID"], "уже не актуальна")
                elif await artDB.update(await self.make_valid_article(art, session), session):
                    pass

        # Видеогалерея
        self.section_id = "69"
        art_inf = await self.get_inf()
        for art in art_inf:
            art_id = art["ID"]

            if "PROPERTY_405" in art:
                pre_section_id = list(art["PROPERTY_405"].values())[0]

                if pre_section_id == "327":
                    art["section_id"] = 42  # Официальные события
                    self.section_id = 42
                elif pre_section_id == "328":
                    art["section_id"] = 52  # Корпоративная жизнь в фото
                    self.section_id = 52

                artDB = ArticleModel(id=art["ID"], section_id=self.section_id)
                if await artDB.need_add(session=session):
                    await self.add(art, session)
                elif await artDB.update(await self.make_valid_article(art, session), session):
                    pass

            else:
                # че делать с уже не актуальными новостями?
                self.section_id = 6
                artDB = ArticleModel(id=art["ID"], section_id=self.section_id)
                if await artDB.need_add(session=session):
                    art["active"] = False
                    await self.add(art, session)
                    print("Запись в фотогалерею", art["NAME"], art["ID"], "уже не актуальна")
                elif await artDB.update(await self.make_valid_article(art, session), session):
                    pass

        # вакансии (приведи друга)
        self.section_id = "67"
        art_inf = await self.get_inf()
        for art in art_inf:
            self.section_id = 111  # потом изменить
            artDB = ArticleModel(id=art["ID"], section_id=self.section_id)
            if await artDB.need_add(session=session):
                await self.add(art, session)
            elif await artDB.update(await self.make_valid_article(art, session), session):
                pass

    async def upload_services(self, session):
        # Корпоративная газета ✔️

        data = [
            {
                "ID": "342022",
                "IBLOCK_ID": "2022",
                "NAME": "№1 (2022)",
                "image": "https://portal.emk.ru/intranet/news/gazeta/img/emk-001.jpg",
                "file": "https://portal.emk.ru/intranet/news/gazeta/pdf/emk-001.pdf",
                "year": "2022",
                "DATE_CREATE": "01.01.2022",
            },
            {
                "ID": "342023",
                "IBLOCK_ID": "2023",
                "NAME": "№2 (2023)",
                "image": "https://portal.emk.ru/intranet/news/gazeta/img/emk-002.jpg",
                "file": "https://portal.emk.ru/intranet/news/gazeta/pdf/emk-002.pdf",
                "year": "2023",
                "DATE_CREATE": "01.01.2023",
            },
            {
                "ID": "342024",
                "IBLOCK_ID": "2024",
                "NAME": "№3 (2024)",
                "image": "https://portal.emk.ru/intranet/news/gazeta/img/emk-003.jpg",
                "file": "https://portal.emk.ru/intranet/news/gazeta/pdf/emk-003.pdf",
                "year": "2024",
                "DATE_CREATE": "01.01.2024",
            }
        ]

        for art in data:
            self.section_id = 34  # потом изменить
            artDB = ArticleModel(id=art["ID"], section_id=self.section_id)
            if await artDB.need_add(session=session):
                await self.add(art, session)
            elif await artDB.update(await self.make_valid_article(art, session), session):
                pass

        # Конкурсы ЭМК 7 секция

        self.section_id = "128"
        competitions_info = await self.get_inf()
        if competitions_info != []:
            for inf in self.logg.progress(competitions_info, "Загрузка 'Конкурсы ЭМК'"):
                # art_id = inf["ID"]
                self.section_id = 71
                art_DB = ArticleModel(id=inf["ID"], section_id=self.section_id)
                if await art_DB.need_add(session=session):
                    await self.add(inf, session)
                elif await art_DB.update(await self.make_valid_article(inf, session), session):
                    pass

        '''самобытные блоки'''
        # полная статика
        # 11 Наша компания -> Наша компания ✔️
        # 12 История компании -> История компании ✔️

        # 110 Техника безопасности -> Техника безопасности ✔️
        # 34 Корпоративная газета ЭМК -> газеты ✔️
        # 41 Гид по предприятиям -> 3D тур ✔️

        # переделки
        # 19 Дни рождения ✔️
        # 21 Подбор оборудования ✔️
        # 22 Поздравительная открытка ♻️
        # 23 ChatGPT ❌
        # 24 Разрешительная документация и сертиффикаты ❌
        # Новые сотрудники ✔️
        # Личный кабинет ✔️
        # Есть Идея ✔️

        # РЕДАКТОРКА

        # новые разделы
        # конфигуратор НПО Регулятор ✔️
        # DeepSeek ❌
        # VCard ✔️
        # система личной эффективности ❌
        # магазин мерча ❌

        # QR-код на САЗ ❌
        # YandexGPT5 + Yandex ART ❌
        # Юбилей САЗ ❌

    async def search_by_id(self, session, user_id: int = None):
        art = await ArticleModel(id=self.id).find_by_id(session)
        files = await File(art_id=int(self.id)).get_files_by_art_id(session)
        art['images'] = []
        art['videos_native'] = []
        art['videos_embed'] = []
        art['documentation'] = []
        if files:
            for file in files:
                # файлы делятся по категориям
                if "image" in file["content_type"] or "jpg" in file["original_name"] or "jpeg" in file[
                    "original_name"] or "png" in file["original_name"]:
                    url = file["file_url"]
                    if art['section_id'] in [42, 51, 52]:
                        preview_link = url.split("/")
                        preview_link[-2] = "compress_image/yowai_mo"
                        url = '/'.join(preview_link)
                    file["file_url"] = f"{DOMAIN}{url}"
                    art['images'].append(file)
                elif "video" in file["content_type"]:
                    url = file["file_url"]
                    file["file_url"] = f"{DOMAIN}{url}"
                    art['videos_native'].append(file)
                elif "link" in file["content_type"]:
                    art['videos_embed'].append(file)
                else:
                    url = file["file_url"]
                    file["file_url"] = f"{DOMAIN}{url}"
                    art['documentation'].append(file)

        prev = await self.get_preview(session)
        art["preview_file_url"] = prev if prev else "https://portal.emk.ru/local/templates/intranet/img/no-user-photo.png"
        # art["preview_file_url"] = await self.get_preview(session)

        if art['section_id'] == 31 or art['section_id'] == 33:
            if 'tags' in art['indirect_data']:
                tags = []
                for tag_id in art['indirect_data']['tags']:
                    tag = {}
                    res = await Tag(id=tag_id).get_tag_by_id(session)
                    tag_name = res.tag_name
                    if tag_name:
                        tag['id'] = tag_id
                        tag['tag_name'] = tag_name
                        tags.append(tag)
                art['indirect_data']['tags'] = tags

        null_list = [17, 19, 111, 112, 14, 18, 25, 54, 55, 53, 56, 7, 71, 34, 175]  # список секций где нет лайков

        if art['section_id'] not in null_list:
            # user_id = await self.get_user_by_session_id(session_id=session_id, session=session)
            if user_id is not None:
                await self.add_art_view(session)
                has_user_liked = await User(id=user_id).has_liked(art_id=self.id, session=session)
                art['reactions'] = has_user_liked

        # обработаем конкурсы эмк где есть лайки, но нет просмотров
        elif art['section_id'] == 71:
            # вызов количества лайков
            del art['indirect_data']['likes_from_b24']
            # user_id = await self.get_user_by_session_id(session_id=session_id, session=session)
            if user_id is not None:
                has_user_liked = await User(id=user_id).has_liked(art_id=self.id, session=session)
                art['reactions'] = has_user_liked

        # магазин мерча
        if art['section_id'] == 56:
            if art['active'] == True:
                size_list = ['s', 'm', 'l', 'xl', 'xxl', 'no_size']
                print(art)
                result = {}
                result['id'] = art['id']
                result['active'] = art['active']
                result['name'] = art['name']
                result['content_text'] = art['content_text']
                result['section_id'] = art['section_id']
                # price = art['indirect_data'].pop('price')
                # photo = art['indirect_data'].pop('preview_file_url')
                # result['price'] = price
                result['indirect_data'] = art['indirect_data']
                sizes_left = dict()
                for size in size_list:
                    if size in art['indirect_data'].keys() and art['indirect_data'][size] is not None:
                        sizes_left[size] = art['indirect_data'][size]
                        art['indirect_data'].pop(size)

                result['indirect_data']['sizes_left'] = sizes_left
                result['indirect_data']['images'] = art['images']

                return result

        return art

    async def delete(self, session):
        # удалить файлы статьи
        try:
            res = await File(art_id=self.id).delete_by_art_id(session=session)
            if res is True:
                return await ArticleModel(id=self.id).remove(session=session)
        except Exception as e:
            return LogsMaker().error_message(f"Возникла ошибка при удалении статьи с id = {self.id}: {e}")

    async def get_preview(self, session):
        files = await File(art_id=int(self.id)).get_files_by_art_id(session=session)
        if files:
            for file in files:
                if file["is_preview"]:
                    url = file["file_url"]

                    # внедряю компрессию
                    if self.section_id == "18":  # отдельный алгоритм для памятки новому сотруднику
                        preview_link = url.split("/")
                        preview_link[-2] = "compress_image/yowai_mo"
                        url = '/'.join(preview_link)
                    # Для баготворительных проектов компрессия не требуется
                    # и для гида по предприятиям

                    elif self.section_id in ["55", "41", "32"]:
                        return f"{DOMAIN}{url}"
                    else:
                        preview_link = url.split("/")
                        preview_link[-2] = "compress_image"
                        # preview_link[-2] = "compress_image/yowai_mo"
                        url = '/'.join(preview_link)

                    return f"{DOMAIN}{url}"

            #Сортируем чтобы файлы были по порядку
            sorted_files = sorted(files, key=lambda x: x['id'], reverse=False)

            # находим первую картинку, если она есть
            for file in sorted_files:
                if "image" in file["content_type"] or "jpg" in file["original_name"] or "jpeg" in file[
                    "original_name"] or "png" in file["original_name"]:
                    # current_num = int(file['name'].split('_')[-1].split('.')[0])
                    # if 1 == current_num:
                    url = file["file_url"]
                    # внедряю компрессию
                    if self.section_id == "18":  # отдельный алгоритм для памятки новому сотруднику
                        preview_link = url.split("/")
                        preview_link[-2] = "compress_image/yowai_mo"
                        url = '/'.join(preview_link)
                    # Для баготворительных проектов компрессия не требуется
                    # и для гида по предприятиям
                    elif self.section_id in ["55", "41", "32"]:
                        return f"{DOMAIN}{url}"
                    else:
                        preview_link = url.split("/")
                        preview_link[-2] = "compress_image"
                        # preview_link[-2] = "compress_image/yowai_mo"
                        url = '/'.join(preview_link)
                    return f"{DOMAIN}{url}"

        return None

    async def find_by_id(self, session):
        art = await ArticleModel(id=self.id).find_by_id(session=session)
        return art

    async def update(self, new_data, session):
        # получаю статью
        art = await ArticleModel(id=self.id).find_by_id(session=session)

        for key in art.keys():
            if key in new_data.keys():
                art[key] = new_data[key]
            elif key == "indirect_data":
                if "indirect_data" in new_data.keys():
                    for subkey in art["indirect_data"].keys():
                        if subkey in new_data["indirect_data"].keys():
                            art["indirect_data"][subkey] = new_data["indirect_data"][subkey]
                else:
                    for subkey in art["indirect_data"].keys():
                        if subkey in new_data.keys():
                            art["indirect_data"][subkey] = new_data[subkey]

        await ArticleModel(id=self.id).update(art, session)

        return True

    async def search_by_section_id(self, session, user_id: int = None):
        if self.section_id == "0":
            main_page = [112, 19, 32, 4, 7, 31, 16, 33, 53, 51]  # 111
            page_view = []

            # user_id = await self.get_user_by_session_id(session_id=session_id, session=session)

            for page in main_page:  # проходимся по каждой секции
                sec = await self.main_page(page, user_id, session)
                page_view.append(sec)
                # page_view[-3]['content'] = [page_view[-2], page_view[-1]]
            # del page_view[-2:]

            return page_view

        elif self.section_id == "19":
            users_bday_info = []
            date_bday = datetime.datetime.now().strftime("%d.%m")
            users = await User().get_birthday_celebrants(date_bday, session)
            return users

        elif self.section_id == "112":
            return await User().get_new_workers(session)



        elif self.section_id == "25" or self.section_id == "175":
            active_articles = []
            result = await ArticleModel(section_id=int(self.section_id)).find_by_section_id(session)
            for res in result:
                if res['active']:
                    self.id = res["id"]

                    # взаимствую логику поиска файлов из метода поиска статей по их id
                    art = await ArticleModel(id=self.id).find_by_id(session)
                    files = await File(art_id=int(self.id)).get_files_by_art_id(session=session)
                    res['images'] = []
                    res['videos_native'] = []
                    res['videos_embed'] = []
                    res['documentation'] = []

                    if files:
                        for file in files:

                            url = file["file_url"]
                            file["file_url"] = f"{DOMAIN}{url}"

                            # файлы делятся по категориям
                            if "image" in file["content_type"] or "jpg" in file["original_name"] or "jpeg" in file[
                                "original_name"] or "png" in file["original_name"]:
                                res['images'].append(file)
                            elif "video" in file["content_type"]:
                                res['videos_native'].append(file)
                            elif "link" in file["content_type"]:
                                res['videos_embed'].append(file)
                            else:

                                res['documentation'].append(file)

                    active_articles.append(res)
            return sorted(active_articles, key=lambda x: x['id'], reverse=True)

        elif self.section_id == "34":
            result = await ArticleModel(section_id=int(self.section_id)).find_by_section_id(session)
            sorted_active_articles = sorted(result, key=lambda x: x['id'], reverse=True)
            return sorted_active_articles

        elif self.section_id == "8":  # Есть Идея
            ideas = await Idea().get_ideas(user_id=user_id, session=session)
            if ideas is not None:
                sorted_active_articles = sorted(ideas, key=lambda x: x['number'], reverse=False)
                return sorted_active_articles
            else:
                return {"err": "Auth Err"}

        # магазин мерча
        elif self.section_id == "56":
            result = []
            res = await ArticleModel(section_id=int(self.section_id)).find_by_section_id(session)
            for re in res:
                if re['active'] == True:
                    images = []
                    self.id = re['id']
                    files = await File(art_id=int(self.id)).get_files_by_art_id(session=session)
                    if files:
                        for file in files:
                            if "image" in file["content_type"] or "jpg" in file["original_name"] or "jpeg" in file[
                                "original_name"] or "png" in file["original_name"]:
                                url = file["file_url"]
                                file["file_url"] = f"{DOMAIN}{url}"
                                images.append(file)

                    # отсюда достать все файлы
                    art_info = {}
                    art_info['id'] = re['id']
                    art_info['section_id'] = re['section_id']
                    art_info['name'] = re['name']

                    if re['indirect_data'] is None:
                        art_info['indirect_data'] = dict()
                    else:
                        art_info['indirect_data'] = re['indirect_data']

                    art_info['indirect_data']['images'] = images

                    result.append(art_info)
            return result

        # конкурсы ЭМК без компрессии
        elif self.section_id == "71":
            active_articles = []
            result = await ArticleModel(section_id=int(self.section_id)).find_by_section_id(session)
            for res in result:
                if res['active']:
                    self.id = res["id"]
                    files = await File(art_id=int(self.id)).get_files_by_art_id(session=session)
                    if files:
                        url = files[0]["file_url"]
                        res['preview_file_url'] = f"{DOMAIN}{url}"
                    else:
                        res['preview_file_url'] = "https://portal.emk.ru/local/templates/intranet/img/no-user-photo.png"
                    # user_id = await self.get_user_by_session_id(session_id=session_id, session=session)
                    if user_id is not None:
                        has_user_liked = await User(id=user_id).has_liked(art_id=self.id, session=session)
                        res['reactions'] = has_user_liked
                    active_articles.append(res)
            sorted_active_articles = sorted(active_articles, key=lambda x: x['id'], reverse=True)
            return sorted_active_articles

        else:
            null_list = [17, 19, 22, 111, 112, 14, 18, 25, 52, 54, 55, 56, 53, 7, 34]  # список секций где нет лайков
            active_articles = []
            result = await ArticleModel(section_id=int(self.section_id)).find_by_section_id(session)
            current_datetime = datetime.datetime.now()
            for res in result:
                if res['active']:
                    if int(self.section_id) in [31, 16, 33]:
                        if res["date_publiction"] is None or (
                                "date_publiction" in res and res["date_publiction"] <= current_datetime):
                            self.id = res["id"]

                            # res["preview_file_url"] = await self.get_preview(session)
                            prev = await self.get_preview(session)
                            res["preview_file_url"] = prev if prev else "https://portal.emk.ru/local/templates/intranet/img/no-user-photo.png"
                            # сюда лайки и просмотры
                            # добавляем лайки и просмотры к статьям раздела. Внимательно добавить в список разделы без лайков
                            # user_id = await self.get_user_by_session_id(session_id=session_id, session=session)
                            if user_id is not None:
                                has_user_liked = await User(id=user_id).has_liked(art_id=self.id, session=session)
                                res['reactions'] = has_user_liked
                        else:
                            continue
                    else:
                        self.id = res["id"]
                        prev = await self.get_preview(session)
                        # res["preview_file_url"] = prev if prev else "https://portal.emk.ru/local/templates/intranet/img/no-user-photo.png"

                        if prev is None:
                            if int(self.section_id) == 32:
                                res["preview_file_url"] = res['indirect_data']['users'][0]['photo_file_url']
                            elif int(self.section_id) == 15:
                                res["preview_file_url"] = prev
                            else:
                                res["preview_file_url"] = "https://portal.emk.ru/local/templates/intranet/img/no-user-photo.png"
                        else:
                            res["preview_file_url"] = prev

                        # сюда лайки и просмотры
                        if int(self.section_id) not in null_list:  # добавляем лайки и просмотры к статьям раздела. Внимательно добавить в список разделы без лайков
                            # user_id = await self.get_user_by_session_id(session_id=session_id, session=session)
                            if user_id is not None:
                                has_user_liked = await User(id=user_id).has_liked(art_id=self.id, session=session)
                                res['reactions'] = has_user_liked

                        # обработаем конкурсы эмк где есть лайки, но нет просмотров
                        # elif res['section_id'] == 7:
                        #     del res['indirect_data']['likes_from_b24']
                        #     # вызов количества лайков
                        #     user_id = self.get_user_by_session_id(session_id=session_id)
                        #     if user_id is not None:
                        #         has_user_liked = User(id=user_id).has_liked(art_id=self.id)
                        #         res['reactions'] = has_user_liked

                    active_articles.append(res)

            if self.section_id == "111" or self.section_id == "14":
                sorted_active_articles = sorted(active_articles, key=lambda x: x['name'], reverse=False)
            # отдельная сортировка Памятки новому сторуднику
            elif self.section_id == "18":
                sorted_active_articles = sorted(active_articles, key=lambda x: int(x['indirect_data']["sort"]),
                                                reverse=False)
            elif self.section_id == "31" or self.section_id == "33":
                sorted_active_articles = sorted(active_articles, key=lambda x: x['date_publiction'], reverse=True)
            else:
                sorted_active_articles = sorted(active_articles, key=lambda x: x['id'], reverse=True)

            return sorted_active_articles

    async def all_serch_by_date(self, session):
        result = await ArticleModel(section_id=self.section_id).find_by_section_id(session)
        sorted_active_articles = sorted(result, key=lambda x: x['id'], reverse=True)
        return sorted_active_articles

    async def main_page(self, section_id, user_id, session):

        # Новые сотрудники
        if section_id == 112:
            img_new_workers = []
            users = await User().get_new_workers(session)
            for user in users:
                user.pop('position')
                user.pop('department')
                user.pop('user_fio')
                img_new_workers.append(user)
            new_workers_view = {
                'id': section_id,
                'type': 'swiper',
                'title': 'Новые сотрудники',
                'images': img_new_workers,
                'href': 'newWorkers',
            }  # словарь-заглушка для будущей секции "новые сотрудники"
            return new_workers_view

        # С днем рождения!
        elif section_id == 19:
            images_for_bday = []
            date_bday = datetime.datetime.now().strftime("%d.%m")
            users = await User().get_birthday_celebrants(date_bday, session)
            for user in users:
                user.pop('position')
                user.pop('department')
                user.pop('user_fio')
                images_for_bday.append(user)

            birthday = {
                'id': section_id,
                'type': 'swiper',
                'title': 'С Днём Рождения!',
                'images': images_for_bday,
                'href': 'birthdays',
            }  # словарь-заглушка для будущей секции "С днем рождения!"
            return birthday

        # Орг развитие
        elif section_id == 32:
            current_datetime = datetime.datetime.now()
            result = []
            articles_in_section = await ArticleModel(section_id=section_id).find_by_section_id(session)
            for values in articles_in_section:
                if values['indirect_data'] is not None and "active_main_page" in values['indirect_data'].keys() and \
                        values['indirect_data']['active_main_page'] == False:
                    continue

                if values["active"] == False:
                    continue

                self.id = values["id"]

                files = await File(art_id=int(self.id)).get_files_by_art_id(session=session)
                image_URL = ""
                for file in files:
                    if "image" in file["content_type"] or "jpg" in file["original_name"] or "jpeg" in file[
                        "original_name"] or "png" in file["original_name"]:
                        url = file["file_url"]
                        image_URL = DOMAIN + url
                if files == [] and values['indirect_data']['users'] != []:
                    image_URL = values['indirect_data']['users'][0]['photo_file_url']
                node = {"id": self.id, "image": image_URL}
                result.append(node)

            #         date_value = [] # список для хранения необходимых данных
            #         date_value.append(values["id"])
            #         date_value.append(values["name"])
            #         date_value.append(values["preview_text"])
            #         date_value.append(values["date_publiction"] if values["date_publiction"] is not None else values["date_creation"])
            #         date_list.append(date_value) # получили список с необходимыми данными
            # # сортируем по дате
            # sorted_data = sorted(date_list, key=lambda x: x[3], reverse=True)

            # for news in sorted_data
            # news_id = sorted_data[0][0]

            # self.id = news_id
            # # image_URL = self.get_preview()
            # files = File(art_id = int(self.id)).get_files_by_art_id()
            # for file in files:
            #     if "image" in file["content_type"] or "jpg" in file["original_name"] or "jpeg" in file["original_name"] or "png" in file["original_name"]:
            #         url = file["file_url"]
            #         image_URL = DOMAIN + url

            second_page = {
                'id': section_id,
                'type': 'swiper',
                'title': 'Организационное развитие',
                "href": "corpNews",
                'images': result
            }
            return second_page

        # предложить идею
        elif section_id == 4:
            idea_block = {
                'id': 4,
                'type': 'swiper',
                'title': 'Предложить идею',
                'images': [{
                    "id": 1,
                    "image": None,
                    "href": "/"
                }],
                'modifiers': ['outline'],
                'href': 'ideasPage'
            }  # словарь-заглушка для будущей секции "Предложить идею"
            return idea_block

        # конкурсы
        elif section_id == 7:
            print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
            articles_in_section = await ArticleModel(section_id=section_id).find_by_section_id(session)
            images = []
            for art in articles_in_section:
                if art["active"] is not False:
                    self.id = art["id"]

                    preview_pict = await self.get_preview(session)
                    # preview_pict = None

                    if preview_pict is None:
                        # image_url = None
                        image_url = "https://portal.emk.ru/local/templates/intranet/img/no-user-photo.png"
                    else:
                        image_url = preview_pict

                    art_img = {
                        "id": self.id,
                        "image": preview_pict,
                        "name": art["name"],
                        "href": art["indirect_data"]["sectionHref"]
                    }
                    images.append(art_img)
            second_page = {
                "id": 7,
                "type": "swiper",
                "title": "Конкурсы ЭМК",
                "images": images
            }

            # print(second_page)

            return second_page


        # Открытые вакансии
        # elif section_id == 111:
        #     emk_competition = {
        #         'id': section_id,
        #         'type': 'singleBlock',
        #         'title': 'Конкурсы ЭМК',
        #         'images': [{
        #             "id": 1,
        #             "image": None,
        #             "href": "vacancies"
        #         }],
        #         '// href': '/'
        #     } # словарь-заглушка для будущей секции "Конкурсы ЭМК"
        #     return emk_competition

        # Актуальные новости
        elif section_id == 31:
            current_datetime = datetime.datetime.now()
            date_list = []  # список для сортировки по дате
            articles_in_section = await ArticleModel(section_id=section_id).find_by_section_id(session)
            for values in articles_in_section:
                if values["active"] is False:
                    pass
                else:
                    date_value = []  # список для хранения необходимых данных
                    if values["date_publiction"] is None or (
                            "date_publiction" in values and values["date_publiction"] <= current_datetime):
                        date_value.append(values["id"])
                        date_value.append(values["name"])
                        date_value.append(values["preview_text"])
                        date_value.append(
                            values["date_publiction"] if values["date_publiction"] is not None else values[
                                "date_creation"])
                        date_list.append(date_value)
                    else:
                        continue

                    # получили список с необходимыми данными
            # сортируем по дате
            sorted_data = sorted(date_list, key=lambda x: x[3], reverse=True)

            second_page = {
                'id': section_id,
                'type': 'section',
                'title': 'Актуальные-новости',
                'href': 'actualArticle',
                'sectionId': 'actualNews',
                'images': []
            }

            business_news = []

            image_url = ''
            for i, row in enumerate(sorted_data):
                if i < 5:
                    news = {}
                    self.id = row[0]
                    preview_pict = await self.get_preview(session)
                    # preview_pict = None
                    if preview_pict is None:
                        # image_url = None
                        image_url = "https://portal.emk.ru/local/templates/intranet/img/no-user-photo.png"
                    else:
                        image_url = preview_pict

                    news['id'] = row[0]
                    news['title'] = row[1]
                    news['date'] = row[3]
                    news['image'] = image_url

                    if user_id is not None:
                        has_user_liked = await User(id=user_id).has_liked(art_id=self.id, session=session)

                        news['reactions'] = has_user_liked
                    business_news.append(news)
            second_page['images'] = business_news
            return second_page

        # Видеоитервью
        elif section_id == 16:
            current_datetime = datetime.datetime.now()
            data_list = []  # список для сортировки по дате
            articles_in_section = await ArticleModel(section_id=section_id).find_by_section_id(session=session)
            for values in articles_in_section:
                if values["active"] is not False:
                    date_value = []  # список для хранения необходимых данных
                    if values["date_publiction"] is None or values["date_publiction"] <= current_datetime:
                        date_value.append(values["id"])
                        date_value.append(values["name"])
                        date_value.append(values["preview_text"])
                        date_value.append(
                            values["date_publiction"] if values["date_publiction"] is not None else values[
                                "date_creation"])
                        data_list.append(date_value)
                    else:
                        continue

                    self.id = values["id"]

                    # data_list.append(data_value) # получили список с необходимыми данными
            # сортируем по дате
            sorted_data = sorted(data_list, key=lambda x: x[0], reverse=True)

            second_page = {
                'id': section_id,
                'type': 'section',
                'title': 'Видеоинтервью',
                'href': 'videoInterview',
                'sectionId': 'videoInterviews',
                'images': []
            }

            interview_news = []

            image_url = ''
            for i, row in enumerate(sorted_data):
                if i < 5:
                    news = {}
                    self.id = row[0]
                    preview_pict = await self.get_preview(session=session)
                    # preview_pict = None
                    if preview_pict is None:
                        # image_url = None
                        image_url = "https://portal.emk.ru/local/templates/intranet/img/no-user-photo.png"
                    else:
                        image_url = preview_pict

                    news['id'] = row[0]
                    news['title'] = row[1]
                    news['description'] = row[2]
                    news['date'] = row[3]
                    news['image'] = image_url
                    # сюда реакции
                    if user_id is not None:
                        has_user_liked = await User(id=user_id).has_liked(art_id=self.id, session=session)

                        news['reactions'] = has_user_liked
                    interview_news.append(news)
            second_page['images'] = interview_news
            return second_page

        # Видеорепортажи
        elif section_id == 33:
            current_datetime = datetime.datetime.now()
            date_list = []  # список для сортировки по дате
            articles_in_section = await ArticleModel(section_id=section_id).find_by_section_id(session=session)
            for values in articles_in_section:
                if values["active"] is False:
                    pass
                else:
                    date_value = []  # список для хранения необходимых данных
                    if values["date_publiction"] is None or values["date_publiction"] <= current_datetime:
                        date_value.append(values["id"])
                        date_value.append(values["name"])
                        date_value.append(values["preview_text"])
                        date_value.append(
                            values["date_publiction"] if values["date_publiction"] is not None else values[
                                "date_creation"])
                        date_list.append(date_value)  # получили список с необходимыми данными
                    else:
                        continue
                        # сортируем по дате
            sorted_data = sorted(date_list, key=lambda x: x[3], reverse=True)

            second_page = {
                'id': section_id,
                'type': 'section',
                'title': 'Видеорепортажи',
                'href': 'videoReport',
                'sectionId': 'videoReports',
                'images': []
            }

            video_news = []

            image_url = ''
            for i, row in enumerate(sorted_data):
                if i < 5:
                    news = {}
                    self.id = row[0]
                    preview_pict = await self.get_preview(session=session)
                    # preview_pict = None
                    if preview_pict is None:
                        # image_url = None
                        image_url = "https://portal.emk.ru/local/templates/intranet/img/no-user-photo.png"
                    else:
                        image_url = preview_pict

                    news['id'] = row[0]
                    news['title'] = row[1]
                    news['description'] = row[2]
                    news['date'] = row[3]
                    news['image'] = image_url
                    # сюда реакции
                    if user_id is not None:
                        has_user_liked = await User(id=user_id).has_liked(art_id=self.id, session=session)

                        news['reactions'] = has_user_liked
                    video_news.append(news)
            second_page['images'] = video_news
            return second_page

        # # микс
        # elif section_id == 9:
        #     second_page = {
        #         "id": 9,
        #         "type": "mixedRowBlock",
        #         "content": []
        #     }
        #     return second_page

        # Афиша
        elif section_id == 53:
            multiple_flag = False
            date_list = []  # список для сортировки по дате
            articles_in_section = await ArticleModel(section_id=section_id).find_by_section_id(session=session)
            for values in articles_in_section:
                if values["active"] is False:
                    pass
                else:
                    date_value = []  # список для хранения необходимых данных
                    date_value.append(values["id"])
                    if "multiple_preview" in values['indirect_data'].keys() and values['indirect_data'][
                        'multiple_preview'] is not None:
                        date_value.append(values['indirect_data']['multiple_preview'])
                    else:
                        date_value.append(None)
                    date_list.append(date_value)  # получили список с необходимыми данными
            # сортируем по дате
            sorted_data = sorted(date_list, key=lambda x: x[0], reverse=True)

            afisha = {
                "id": 53,
                'type': "swiper",
                'title': "Афиша",
                'href': 'eventAnnounces',
                'images': []
            }
            image_url = ''
            afisha_news = []
            for i, row in enumerate(sorted_data):
                if i < 5:
                    news = {}
                    self.id = row[0]
                    if row[1] == True:
                        files = await File(art_id=int(self.id)).get_files_by_art_id(session)
                        if files:
                            preview_pict = []
                            for file in files:
                                # файлы делятся по категориям
                                if "image" in file["content_type"] or "jpg" in file["original_name"] or "jpeg" in file[
                                    "original_name"] or "png" in file["original_name"]:
                                    url = file["file_url"]
                                    preview_link = url.split("/")
                                    preview_link[-2] = "compress_image"
                                    url = '/'.join(preview_link)
                                    preview_pict.append(f"{DOMAIN}{url}")
                        else:
                            preview_pict = None
                    else:
                        preview_pict = await self.get_preview(session=session)
                    # preview_pict = None
                    if preview_pict is None:
                        # image_url = None
                        image_url = "https://portal.emk.ru/local/templates/intranet/img/no-user-photo.png"
                    else:
                        image_url = preview_pict

                    news['id'] = row[0]
                    news['image'] = image_url
                    afisha_news.append(news)

            afisha['images'] = afisha_news

            return afisha

        # Корпоративные события
        elif section_id == 51:
            date_list = []  # список для сортировки по дате
            articles_in_section = await ArticleModel(section_id=section_id).find_by_section_id(session=session)
            for values in articles_in_section:
                if values["active"] is False:
                    pass
                else:
                    date_value = []  # список для хранения необходимых данных
                    date_value.append(values["id"])
                    date_value.append(values["name"])
                    date_value.append(values["preview_text"])
                    date_value.append(
                        values["date_publiction"] if values["date_publiction"] is not None else values["date_creation"])
                    date_list.append(date_value)  # получили список с необходимыми данными
            # сортируем по дате
            sorted_data = sorted(date_list, key=lambda x: x[0], reverse=True)

            corpevents = {
                'id': 51,
                'type': "section",
                'title': "Корпоративные события",
                'href': 'corpEvent',
                'sectionId': 'corpEvents',
                'images': []
            }
            image_url = ''
            corpevents_news = []
            for i, row in enumerate(sorted_data):
                if i < 5:
                    news = {}
                    self.id = row[0]
                    preview_pict = await self.get_preview(session=session)
                    # preview_pict = None
                    if preview_pict is None:
                        # image_url = None
                        image_url = "https://portal.emk.ru/local/templates/intranet/img/no-user-photo.png"
                    else:
                        image_url = preview_pict

                    news['id'] = row[0]
                    news['title'] = row[1]
                    news['description'] = row[2]
                    news['date'] = row[3]
                    news['image'] = image_url
                    # сюда реакции
                    if user_id is not None:
                        has_user_liked = await User(id=user_id).has_liked(art_id=self.id, session=session)

                        news['reactions'] = has_user_liked
                    corpevents_news.append(news)

            corpevents['images'] = corpevents_news
            return corpevents

    # лайки
    # если раскомиттить - вызовет ошибку, нет глобальной сессии внутри функции get_likes_count()
    # async def get_all_likes(self):
    #     return await LikesModel(art_id=self.id).get_likes_count()

    async def add_like(self, user_id, session):
        # user_id = await self.get_user_by_session_id(session_id=session_id, session=session)
        if user_id is not None:
            return await LikesModel(user_id=user_id, art_id=self.id).add_or_remove_like(session=session)
        return {"err": "Auth Err"}

    async def has_user_liked(self, user_id, session):
        # user_id = await self.get_user_by_session_id(session_id=session_id, session=session)
        if user_id is not None:
            return await LikesModel(user_id=user_id, art_id=self.id).has_liked(session=session)
        return {"err": "Auth Err"}

    # # просмотры
    # async def get_art_views(self):
    #     return await ViewsModel(art_id=self.id).get_art_viewes()

    async def add_art_view(self, session):
        return await ViewsModel(art_id=self.id).add_art_view(session=session)

    # дамп данных по лайкам из Б24
    async def upload_likes(self, session):
        result = []
        articles_info = await ArticleModel().all(session=session)
        null_list = [17, 19, 111, 112, 14, 18, 25, 54, 55, 53, 7, 34]  # список секций где нет лайков
        for inf in articles_info:
            if inf['section_id'] not in null_list:
                # конкурсы ЭМК
                if inf['section_id'] == 71:
                    if 'likes_from_b24' in inf['indirect_data'] and inf['indirect_data']['likes_from_b24'] is not None:
                        for user_id in inf['indirect_data']['likes_from_b24']:
                            user_exist = await User(int(user_id)).search_by_id_all(session=session)
                            if isinstance(user_exist, types.CoroutineType) or user_exist is None:
                                continue
                            else:
                                has_usr_liked = await LikesModel(user_id=int(user_id), art_id=int(inf['id'])).has_liked(
                                    session=session)
                                if has_usr_liked['likes']['likedByMe']:
                                    continue
                                else:
                                    await LikesModel(user_id=int(user_id), art_id=int(inf['id'])).add_or_remove_like(
                                        session=session)
                        await ArticleModel(id=int(inf['id'])).remove_b24_likes(session=session)

                # все остальное
                else:
                    likes_info = B24().get_likes_views(inf['id'])
                    if likes_info != "Not found" and 'VOTES' in likes_info.keys():
                        for vote in likes_info['VOTES']:
                            # проверяем есть ли такие юзеры в бд
                            user_exist = await User(vote['USER_ID']).search_by_id_all(session=session)
                            if isinstance(user_exist, types.CoroutineType) or user_exist is None:
                                continue
                            else:
                                await LikesModel(user_id=vote['USER_ID'], art_id=inf['id']).add_like_from_b24(
                                    vote['CREATED_'], session=session)

                        # удаляем тех, кто убрал лайк
                        b24_likers = [i['USER_ID'] for i in likes_info['VOTES']]
                        article_likers = await LikesModel(art_id=inf['id']).get_article_likers(session=session)
                        for usr in article_likers:
                            if usr not in b24_likers:
                                await LikesModel(user_id=usr, art_id=inf['id']).add_or_remove_like(session=session)
                            else:
                                pass

        return {"status": True}

    # дамп данных по просмотрам из Б24
    async def upload_views(self, session):
        result = []
        articles_info = await ArticleModel().all(session=session)
        null_list = [17, 19, 111, 112, 14, 18, 25, 54, 55, 53, 7, 71, 34]  # список секций где нет лайков
        for inf in articles_info:
            if inf['section_id'] not in null_list:

                likes_info = B24().get_likes_views(inf['id'])
                if likes_info != "Not found" and 'VIEWS' in likes_info.keys():
                    VM = ViewsModel()
                    VM.views_count = likes_info['VIEWS']
                    VM.art_id = inf['id']
                    await VM.add_view_b24(session=session)

                    print(likes_info["ID"],
                          "добавил просмотры")  # !!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!!

        return {"status": True}

    # дамп данных в эластик
    async def dump_articles_data_es(self, session):
        return await ArticleSearchModel().dump(session=session)

    # для статистики лайки и просмотры
    async def get_article_likers(self, session):
        return LikesModel(art_id=self.id).get_article_likers(session=session)

    async def get_popular_articles(self, limit, session):
        return LikesModel().get_popular_articles(limit=limit, session=session)

    async def get_recent_popular_articles(self, days, limit, session):
        return LikesModel().get_recent_popular_articles(days=days, limit=limit, session=session)

    async def get_user_by_session_id(self, user_id, session):
        from src.services.Auth import AuthService
        #user = await AuthService().get_user_info(session_id)

        if user_id is not None:
            #user_id = user["ID"]

            # получить и вывести его id
            usr = User()
            usr.id = int(user_id)
            user_inf = await usr.search_by_id(session=session)
            if user_inf is not None and "id" in user_inf.keys():
                return user_inf["id"]
        return None

    async def search_articles_by_tags(self, tag_id, session, user_id: int = None):
        #user_id = await self.get_user_by_session_id(session_id=session_id, session=session)
        result = await Tag(id=tag_id).get_articles_by_tag_id(self.section_id, session=session)
        if result != []:
            sorted_active_articles = sorted(result, key=lambda x: x.date_publiction, reverse=True)
            res = []
            for art in sorted_active_articles:
                self.id = art.id
                art = art.__dict__

                prev = await self.get_preview(session)
                art[
                    'preview_file_url'] = prev if prev else "https://portal.emk.ru/local/templates/intranet/img/no-user-photo.png"

                if user_id is not None:
                    has_user_liked = await User(id=user_id).has_liked(art_id=self.id, session=session)

                    art['reactions'] = has_user_liked
                res.append(art)

            return res
        else:
            return result

    async def set_tag_to_art_id(self, tag_id, session):
        return await Tag(id=tag_id, art_id=self.id).set_tag_to_art_id(session=session)

    async def remove_tag_from_art_id(self, tag_id, session):
        return await Tag(id=tag_id, art_id=self.id).remove_tag_from_art_id(session=session)

    async def check_user_root(self, user_id, session):
        from ..services.Fieldsvisions import Visions
        return await Visions(art_id=self.id, user_id=user_id).check_user_root(session=session)

    async def update_art_el_index(self, data):
        return await ArticleSearchModel().update_art_el_index(article_data=data, session=session)

    async def get_token_by_uuid(self, session, user_id):
        from src.services.Roots import Roots
        roots_token = await Roots(user_uuid=user_id).get_token_by_uuid(session=session)
        return roots_token

    async def make_event_users_excel(self, session, data):
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Список участников"

            # Запись данных
            ws['A1'] = 'ФИО'
            ws['B1'] = 'EMAIL'
            ws['C1'] = 'Телефон'
            ws['D1'] = 'Внутр. номер'
            ws['E1'] = 'Должность'
            ws['F1'] = 'Дирекция/Завод'
            ws['G1'] = 'Подразделение'
            ws['H1'] = 'Местоположение'

            for i, user in enumerate(data, start=2):
                user_inf = await User(id=user['id']).search_by_id_all(session)

                indirect_data = user_inf.get("indirect_data", {})

                if "name" in user_inf and "last_name" in user_inf and "second_name" in user_inf: ws[
                    f'A{i}'] = f'{user_inf["name"]} {user_inf["last_name"]} {user_inf["second_name"]}'

                if "email" in user_inf: ws[f'B{i}'] = f'{user_inf["email"]}'
                if "personal_mobile" in user_inf: ws[f'C{i}'] = f'{user_inf["personal_mobile"]}'
                if "uf_phone_inner" in user_inf: ws[f'D{i}'] = f'{user_inf["uf_phone_inner"]}'
                if "work_position" in indirect_data:
                    ws[f'E{i}'] = f'{indirect_data["work_position"]}'
                if "uf_department" in indirect_data and "uf_usr_1696592324977" in indirect_data:
                    ws[f'F{i}'] = ", ".join(indirect_data["uf_usr_1696592324977"]) + "/" + ", ".join(
                        indirect_data["uf_department"])
                if "uf_usr_1705744824758" in indirect_data:
                    ws[f'G{i}'] = " ".join(indirect_data["uf_usr_1705744824758"])
                if "personal_city" in user_inf: ws[f'H{i}'] = f'{user_inf["personal_city"]}'

            excel_buffer = io.BytesIO()
            wb.save(excel_buffer)
            excel_buffer.seek(0)

            # Сохранение
            return excel_buffer
        except Exception as e:
            return LogsMaker().error_message(f'Произошла ошибка при создании файла excel make_event_users_excel: {e}')

    async def delete_duplicate_video(self, session, user_id):
        VID_SEC = [31, 33, 42, 51, 52] # СПИСОК СЕКЦИЙ ГДЕ ЕСТЬ ВИДОСЫ
        # ПРОЙТИСЬ ПО СПИСКУ СЕКЦИЙ
        # ПРОЙТИСЬ ПО КАЖДОЙ СТАТЬЕ И СДЕЛАТЬ ЗАПРОС НА ФАЙЛЫ СТАТЬИ
        # ЕСЛИ ЕСТЬ LINK ТО ОСТАВИТЬ ОДИН 
        try:
            for sec in VID_SEC:
                self.section_id = sec
                articles = self.search_by_section_id(session=session, user_id=user_id)
                for art in articles:
                    art_id = art['id']
                    files = await File(art_id=int(art_id)).get_files_by_art_id(session)
                    videos_embed = []
                    if files:
                        for file in files:
                            # файлы делятся по категориям
                            if "link" in file["content_type"]:
                                videos_embed.append(file)
                    if videos_embed:
                        seen_url = []
                        for vid in videos_embed:
                            if vid['b24_url'] in seen_url:
                                #удаляем
                                File(id=vid['id']).editor_del_file(session)
                            else:
                                seen_url.append(vid['b24_url'])
            return True
        except Exception as e:
            return {"error": str(e)}

                    
    async def get_all(self, session):
        articles_info = await ArticleModel().all(session=session)
        return articles_info



# Получить данные инфоблока из Б24
@article_router.get("/infoblock/{ID}", tags=["Статьи", "Битрикс24"])
async def test(ID):
    """
    
    ## Метод `lists.element.get`

    > Метода вызывается один раз, в момент запуска сервиса для загрузки данных с прошлой версии сайта

    Получает элементы информационного блока (списка) из Битрикс24 по его ID через API метод `lists.element.get`.

    ### Входные параметры
    | Параметр | Тип | Описание | Обязательный |
    |----------|-----|----------|--------------|
    | `id` | integer/string | ID информационного блока (IBLOCK_ID) в Битрикс24 | Да |

    ### Возвращаемые данные
    Возвращает список элементов информационного блока. Каждый элемент содержит следующие поля:
    - `ID` (int/string) — уникальный идентификатор элемента
    - `NAME` (string) — название элемента
    - `IBLOCK_ID` (int/string) — ID информационного блока
    - `IBLOCK_SECTION_ID` (int/string) — ID раздела (если есть)
    - `ACTIVE` (string) — активность элемента (Y/N)
    - `DATE_CREATE` (string) — дата создания
    - `CREATED_BY` (int/string) — ID создателя
    - `PROPERTY_*` — пользовательские свойства информационного блока

    ### Пример ответа
    ```json
    [
    {
        "ID": "123",
        "NAME": "Статья о интеграции",
        "IBLOCK_ID": "45",
        "IBLOCK_SECTION_ID": null,
        "ACTIVE": "Y",
        "DATE_CREATE": "2024-01-15T10:30:00+03:00",
        "CREATED_BY": "789",
        "PROPERTY_TAGS": "интеграция, API, документация",
        "PROPERTY_AUTHOR": "Иванов И.И.",
        "PROPERTY_CONTENT": "Текст статьи..."
    },
    {
        "ID": "124",
        "NAME": "Руководство пользователя",
        "IBLOCK_ID": "45",
        "IBLOCK_SECTION_ID": "5",
        "ACTIVE": "Y",
        "DATE_CREATE": "2024-01-16T14:45:00+03:00",
        "CREATED_BY": "456",
        "PROPERTY_TAGS": "документация, руководство",
        "PROPERTY_AUTHOR": "Петров П.П.",
        "PROPERTY_CONTENT": "Руководство пользователя системы..."
    }
    ]
    """
    return await Article(section_id=ID).get_inf()


# загрузить статьи из иноблоков Битрикса
@article_router.put("", tags=["Статьи"])
# , description="""
# ## Метод `get_file(id, inf_id)`

# Получает информацию о прикрепленном файле из диска Битрикс24 по его ID и ID сущности.

# ### Входные параметры
# | Параметр | Тип | Описание | Обязательный |
# |----------|-----|----------|--------------|
# | `id` | integer/string | ID прикрепленного файла в Битрикс24 | Да |
# | `inf_id` | integer/string | ID сущности (элемента инфоблока), к которой прикреплен файл | Да |

# ### Возвращаемые данные
# Возвращает словарь с информацией о файле. Ключевые поля:
# - `ID` (string) — уникальный идентификатор файла
# - `NAME` (string) — имя файла в системе
# - `FILE_NAME` (string) — оригинальное имя файла
# - `SIZE` (string) — размер файла в байтах
# - `DOWNLOAD_URL` (string) — URL для скачивания файла
# - `SRC` (string) — путь к файлу на сервере Битрикс24

# ### Пример ответа
# ```json
# {
#     "ID": "789",
#     "NAME": "document.pdf",
#     "FILE_NAME": "technical_specification.pdf",
#     "SIZE": "1048576",
#     "DOWNLOAD_URL": "https://portal.emk.ru/download/789/",
#     "SRC": "/upload/iblock/document.pdf"
# }
# """
# )
async def upload_articles(session: AsyncSession = Depends(get_async_db)):
    return await Article().uplod(session)


# найти статью по id
@article_router.get("/find_by_ID/{ID}", tags=["Статьи"])
async def get_article(ID: int, request: Request, session: AsyncSession = Depends(get_async_db)):
    user_id = ""
    token = request.cookies.get("user_id")
    if token is None:
        token = request.cookies.get("user_id")
        if token is not None:
            user_id = token
    else:
        user_id = token
    art = Article()
    art.id = ID
    return await art.search_by_id(user_id=user_id, session=session)


# найти статьи раздела
@article_router.get("/find_by/{section_id}", tags=["Статьи"])
async def get_articles(section_id, request: Request, session: AsyncSession = Depends(get_async_db)):
    user_id = ""
    token = request.cookies.get("user_id")
    if token is None:
        token = request.cookies.get("user_id")
        if token is not None:
            user_id = token
    else:
        user_id = token

    if user_id == "undefind":
        return {"err": "Undefined section_id!"}
    else:
        art = Article()
        art.section_id = section_id
        return await art.search_by_section_id(user_id=user_id, session=session)


@article_router.put("/add_or_remove_like/{article_id}", tags=["Статьи"])
async def add_or_remove_like(article_id, request: Request, session: AsyncSession = Depends(get_async_db)):
    user_id = ""
    token = request.cookies.get("user_id")
    if token is None:
        token = request.cookies.get("user-id")
        if token is not None:
            user_id = token
    else:
        user_id = token
    art = Article()
    art.id = article_id
    return await art.add_like(user_id=user_id, session=session)


@article_router.get("/has_user_liked/{article_id}", tags=["Статьи"])
async def has_user_liked(article_id, request: Request, session: AsyncSession = Depends(get_async_db)):
    user_id = ""
    token = request.cookies.get("user_id")
    if token is None:
        token = request.cookies.get("user-id")
        if token is not None:
            user_id = token
    else:
        user_id = token
    art = Article()
    art.id = article_id
    return await art.has_user_liked(user_id=user_id, session=session)


# поиск по статьям еластик
@article_router.get("/search/full_search_art/{keyword}", tags=["Статьи"])
async def elastic_search(keyword: str):
    return await ArticleSearchModel().elasticsearch_article(key_word=keyword)


# выгрузка данных по лайкам в Б24
@article_router.put("/put_b24_likes", tags=["Статьи", "Битрикс24"], description="""
## Метод `getLikes.php`

> Метода вызывается один раз, в момент запуска сервиса для загрузки данных с прошлой версии сайта

Получает информацию о лайках (и просмотрах) для конкретной статьи через специальный PHP endpoint портала Битрикс24.

### Входные параметры
| Параметр | Тип | Описание | Обязательный |
|----------|-----|----------|--------------|
| `art_id` | integer/string | ID статьи в системе Битрикс24 | Да |

### Возвращаемые данные
Возвращает словарь с информацией о лайках и просмотрах статьи. Возможные форматы ответа:

**Успешный ответ (статья найдена):**
```json
{
    "VOTES": [
        {
            "ID": "123",
            "USER_ID": "456",
            "CREATED_": "2024-01-15 10:30:00",
            "VOTE_VALUE": "1"
        },
        {
            "ID": "124",
            "USER_ID": "789",
            "CREATED_": "2024-01-15 11:45:00",
            "VOTE_VALUE": "1"
        }
    ],
    "VIEWS": "150"
}
"""
)
async def put_b24_likes(session: AsyncSession = Depends(get_async_db)):
    return await Article().upload_likes(session)


@article_router.put("/put_b24_views", tags=["Статьи", "Битрикс24"], description="""
## Метод `getLikes.php`

> Метода вызывается один раз, в момент запуска сервиса для загрузки данных с прошлой версии сайта

Получает информацию о просмотрах (и лайках) для конкретной статьи через специальный PHP endpoint портала Битрикс24.

### Входные параметры
| Параметр | Тип | Описание | Обязательный |
|----------|-----|----------|--------------|
| `art_id` | integer/string | ID статьи в системе Битрикс24 | Да |

### Возвращаемые данные
Возвращает словарь с информацией о просмотрах статьи. Возможные форматы ответа:

**Успешный ответ (статья найдена):**
```json
{
    "ID": "12345",
    "VIEWS": "150",
    "VOTES": [
        {
            "ID": "1",
            "USER_ID": "456",
            "CREATED_": "2024-01-15 10:30:00",
            "VOTE_VALUE": "1"
        }
    ]
}
""")
async def put_b24_views(session: AsyncSession = Depends(get_async_db)):
    return await Article().upload_views(session)


# лайки и просмотры для статистики
@article_router.get("/get_article_likers/{ID}", tags=["Статьи"])
async def get_article_likers(ID: int, session: AsyncSession = Depends(get_async_db)):
    art = Article()
    art.id = ID
    return await art.get_article_likers(session)


@article_router.get("/get_popular_articles/{limit}", tags=["Статьи"])
async def get_popular_articles(limit: int, session: AsyncSession = Depends(get_async_db)):
    return await Article().get_popular_articles(limit=limit, session=session)


@article_router.get("/get_recent_popular_articles/{days}/{limit}", tags=["Статьи"])
async def get_recent_popular_articles(days: int, limit: int, session: AsyncSession = Depends(get_async_db)):
    return await Article().get_recent_popular_articles(days=days, limit=limit, session=session)


@article_router.get("/get_articles_by_tag_id/{section_id}/{tag_id}", tags=["Статьи"])
async def get_articles_by_tag_id(section_id: int, tag_id: int, request: Request,
                                 session: AsyncSession = Depends(get_async_db)):
    user_id = ""
    token = request.cookies.get("user_id")
    if token is None:
        token = request.cookies.get("user-id")
        if token is not None:
            user_id = token
    else:
        user_id = token
    art = Article()
    art.section_id = section_id
    return await art.search_articles_by_tags(tag_id=tag_id, user_id=user_id, session=session)


@article_router.put("/set_tag_to_art_id/{tag_id}/{art_id}", tags=["Статьи"])
async def set_tag_to_art_id(art_id: int, tag_id: int, session: AsyncSession = Depends(get_async_db)):
    return await Article(id=art_id).set_tag_to_art_id(tag_id=tag_id, session=session)


@article_router.delete("/remove_tag_from_art_id/{tag_id}/{art_id}", tags=["Статьи"])
async def remove_tag_from_art_id(art_id: int, tag_id: int, session: AsyncSession = Depends(get_async_db)):
    return await Article(id=art_id).remove_tag_from_art_id(tag_id=tag_id, session=session)


@article_router.post("/make_event_users_excel", summary="Скачать Excel со всеми участниками мероприятия", tags=["Статьи"])
async def make_users_excel_list(request: Request, data: list = Body(), session: AsyncSession = Depends(get_async_db)):
    user_id = None
    token = request.cookies.get("user_id")
    if token is None:
        token = request.cookies.get("user-id")
        if token is not None:
            user_id = token
    else:
        user_id = token

    art_class = Article()
    # user_id = await art_class.get_user_by_session_id(user_id=user_id, session=session)
    roots_token = await art_class.get_token_by_uuid(session=session, user_id=user_id)
    if roots_token:
        if ("EditorAdmin" in roots_token.keys() and roots_token["EditorAdmin"] is True) or (
                "EditorModer" in roots_token.keys() and 53 in roots_token["EditorModer"]):
            excel_buffer = await art_class.make_event_users_excel(session=session, data=data)
            return StreamingResponse(excel_buffer,
                                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                    headers={"Content-Disposition": "attachment; filename=participants.xlsx"})

    return LogsMaker().warning_message(f"Недостаточно прав для скачивания Экселя")

# загрузить дату в эластик
@article_router.put("/elastic_data")
async def upload_articles_to_es(session: AsyncSession = Depends(get_async_db)):
    return await ArticleSearchModel().dump(session)

@article_router.get("/give_double")
async def dubli(request: Request, session: AsyncSession = Depends(get_async_db)):
    user_id = ""
    token = request.cookies.get("user_id")
    if token is None:
        token = request.cookies.get("user_id")
        if token is not None:
            user_id = token
    else:
        user_id = token

    if user_id == "undefind":
        return {"err": "Undefined section_id!"}
    art_class = await Article().delete_duplicate_video(session=session, user_id=user_id)
    #по всем сатьями
    return art_class
    #где есть videos_embed
    #если есть повтор 
    #показать
    #удалить пока они не кончаться



# #найти статьи раздела по названию
# @article_router.post("/search/title/{title}")
# async def search_articles_by_title(title): # data = Body()
#     return ArticleSearchModel().search_by_title(title)

# #найти статьи раздела по заголовку
# @article_router.post("/search/preview/{preview}")
# async def search_articles_by_preview(preview): # data = Body()
#     return ArticleSearchModel().search_by_preview(preview)

# #найти статьи раздела по тексту
# @article_router.post("/search/text/{text}")
# async def search_articles_by_text(text): # data = Body()
#     return ArticleSearchModel().search_by_text(text)

# лайки и просмотры для статистики
# @article_router.get("/get_all_likes/{ID}")
# async def get_all_likes(ID: int):
#     return Article(id = ID).get_all_likes()

# @article_router.get("/get_viewers/{ID}")
# async def get_viewers(ID: int):
#     return Article(id = ID).get_art_views()

